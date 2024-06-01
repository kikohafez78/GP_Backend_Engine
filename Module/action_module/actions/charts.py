from xlwings import constants as win32c
from constants import constants
import win32api
import win32com.client as win32
from typing import Any, Optional, List
from openpyxl.utils import get_column_letter
import os
import sys
script_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.append(script_dir)
import itertools

class Charts_App():
    def __init__(self, app = 'excel', api_doc = None) -> None:
        super().__init__()
        self.appName = app
        self.__excel = None
        self.__currentWB = None

        if api_doc is not None:
            for key, value in api_doc.items():
                if value.get('display') is not None:
                    setattr(self, value['display'], self.__getattribute__(key))

    @property
    def activeAPP(self):
        if not self.__excel:
            try:
                self.__excel = win32.Dispatch('Excel.Application') if self.appName == 'excel' else win32.Dispatch('ket.Application')
                self.__excel.DisplayAlerts = False
                self.__excel.Visible = False
            except:
                raise Exception('{} is not running.'.format(self.appName))
        return self.__excel
    
    @property
    def activeWB(self):
        if self.__currentWB is not None:
            return self.__currentWB
        return self.activeAPP.ActiveWorkbook
    
    @activeWB.setter
    def activeWB(self, wb):
        self.__currentWB = wb
    
    @property
    def activeWS(self):
        return self.activeWB.ActiveSheet
    
    def toRangeUsingSheet(self, source: str):
        if '!' in source:
            sheet_name, source = source.split('!')
            sheet_name = sheet_name.strip("'") 
            if sheet_name not in [sheet.Name for sheet in self.activeWB.Worksheets]:
                raise ValueError(f'Sheet {sheet_name} does not exist.')
            sheet = self.activeWB.Sheets(sheet_name)
        else:
            raise Exception('The range must contain a sheet name.')
        
        if source.isdigit():
            return sheet.Rows(source)
        elif source.isalpha():
            return sheet.Columns(source)
        elif ':' in source or source.isalnum():
            return sheet.Range(source)
        
    def toRange(self, sheet,  source: str):
        if source.isdigit():
            return sheet.Rows(source)
        elif source.isalpha():
            return sheet.Columns(source)
        elif ':' in source or source.isalnum():
            return sheet.Range(source)
    
    def OpenWorkbook(self, path: str) -> None:
        self.__currentWB = self.activeAPP.Workbooks.Open(os.path.abspath(path))

    def SaveWorkbook(self, path: str) -> None:
        self.activeWB.SaveAs(os.path.abspath(path))
    
    def closeWorkBook(self) -> None:
        self.activeWB.Close()

    def Save(self) -> None:
        # self.activeWB.Save()
        pass
        
    def CreateChart(self, src_wb_path: str, source: str, destSheet: str, chartType: str, chartName: str, XField: int = None, YField: List[int] = [], output_wb_path: str = ""):
        if output_wb_path == None:
            output_wb_path = src_wb_path
        self.OpenWorkbook(src_wb_path)
        if destSheet not in [sheet.Name for sheet in self.activeWB.Worksheets]:
                raise ValueError(f'Sheet {destSheet} does not exist.')
        sheet = self.activeWB.Sheets(destSheet)
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    raise ValueError(f'The chart name {chartName} already exists.')
        if destSheet not in [sheet.Name for sheet in self.activeWB.Worksheets]:
            raise ValueError(f'Sheet {destSheet} does not exist.')
        sheet = self.activeWB.Worksheets(destSheet)
        dataRange = self.toRange(sheet, source)
        destRange = self.GetBlankArea(destSheet)
        chart = sheet.ChartObjects().Add(destRange.Left, destRange.Top, 350, 200).Chart
        
        if chartType not in constants.ChartType:
            raise ValueError(f'Chart type {chartType} is not supported!')
        
        chart.chart_type = constants.ChartType[chartType]
        
        if 'pie' in chartType.lower():
            chart.SetSourceData(dataRange)
        else:
            if not XField:
                XField = 1
            XFieldRange = dataRange.Parent.Range(dataRange.Cells(2, XField), dataRange.Cells(dataRange.Rows.Count, XField))
            if not YField:
                YField = [i for i in range(1, dataRange.Columns.Count + 1) if i != XField]
            for i in YField:
                series = chart.SeriesCollection().NewSeries()
                series.XValues = XFieldRange
                series.Values = dataRange.Parent.Range(dataRange.Cells(2, i), dataRange.Cells(dataRange.Rows.Count, i))
                series.Name = dataRange.Cells(1, i)
            try:
                chart.Axes(constants.AxisType['x']).CategoryNames = XFieldRange
            except:
                pass
        chart.Parent.Name = chartName
        self.SaveWorkbook(output_wb_path)
        self.closeWorkBook()

    def SetChartTrendline(self, src_wb_path: str, chartName: str, trendlineType: List[str], DisplayEquation: Optional[bool] = None,
                          DisplayRSquared: Optional[bool] = None, output_wb_path: str = None) -> None:
        if output_wb_path == None:
            output_wb_path = src_wb_path
        self.OpenWorkbook(src_wb_path)
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        trendlineType = itertools.cycle(trendlineType)
        for series in chart.SeriesCollection():
            for trendline in series.Trendlines():
                trendline.Delete()
            series.Trendlines().Add(
                constants.TrendlineType[next(trendlineType)],
                DisplayEquation=DisplayEquation,
                DisplayRSquared=DisplayRSquared
            )
        self.SaveWorkbook(output_wb_path)
        self.closeWorkBook()

    def SetChartTitle(self,src_wb_path: str, chartName: str, title: str, fontSize: Optional[float] = None, 
                        bold: bool = None, color: Optional[int] = None, output_wb_path: str = None) -> None:
        if output_wb_path == None:
            output_wb_path = src_wb_path
        self.OpenWorkbook(src_wb_path)
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        chart.HasTitle = True
        chart.ChartTitle.Text = title
        if fontSize:
            chart.ChartTitle.Font.Size = fontSize
        if color:
            chart.ChartTitle.Font.ColorIndex = {
                'black': 1,
                'white': 2,
                'red': 3,
                'green': 4,
                'blue': 5,
                'yellow': 6,
                'magenta': 7,
                'cyan': 8,
                'dark red': 9,
                'dark green': 10
            }[color]
        if not bold is None:
            chart.ChartTitle.Font.Bold = bold
        self.SaveWorkbook(output_wb_path)
        self.closeWorkBook()

    def SetChartHasAxis(self, src_wb_path : str, chartName: str, axis: str, hasAxis: bool, output_wb_path: str) -> None:
        if output_wb_path == None:
            output_wb_path = src_wb_path
        self.OpenWorkbook(src_wb_path)
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        axis = constants.AxisType[axis]
        axisGroup = constants.AxisGroup['primary']
        chart.SetHasAxis(axis, axisGroup, hasAxis)
        self.SaveWorkbook(output_wb_path)
        self.closeWorkBook()
        

    def SetChartAxis(self, src_wb_path: str, chartName: str, axis: str, title: Optional[str] = None, 
                        labelOrientation: Optional[str] = None, maxValue: Optional[float] = None,
                        miniValue: Optional[float] = None, output_wb_path: str = None) -> None:
        if output_wb_path == None:
            output_wb_path = src_wb_path
        self.OpenWorkbook(src_wb_path)
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        if axis in ['x', 'X']:
            axis = win32c.AxisType.xlCategory
        elif axis in ['y', 'Y']:
            axis = win32c.AxisType.xlValue
        elif axis in ['z', 'Z']:
            axis = win32c.AxisType.xlSeriesAxis
        else:
            print('Not support axes type')
            return
        chartAxes = chart.Axes(axis)
        if title:
            chartAxes.HasTitle = True
            chartAxes.AxisTitle.Text = title
        if labelOrientation:
            labelOrientation = {
                'upward': win32c.Orientation.xlUpward,
                'downward': win32c.Orientation.xlDownward,
                'horizontal': win32c.Orientation.xlHorizontal,
                'vertical': win32c.Orientation.xlVertical
                }[labelOrientation]
            chartAxes.TickLabels.Orientation = labelOrientation
        if maxValue:
            chartAxes.MaximumScale = maxValue
        if miniValue:
            chartAxes.MinimumScale = miniValue
        self.SaveWorkbook(output_wb_path)
        self.closeWorkBook()

    def SetChartLegend(self, src_wb_path: str, chartName: str, position: Optional[str] = None, fontSize: Optional[str] = None,
                        seriesName: Optional[list] = [], output_wb_path: str = None) -> None:
        # find the chart
        if output_wb_path == None:
            output_wb_path = src_wb_path
        self.OpenWorkbook(src_wb_path)
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        chart.HasLegend = True
        
        if position and position != 'None':
            # For legent position enumeration, refer to https://learn.microsoft.com/en-us/dotnet/api/microsoft.office.interop.word.xllegendposition?view=word-pia
            position = {
                'bottom': win32c.LegendPosition.xlLegendPositionBottom,
                'corner': win32c.LegendPosition.xlLegendPositionCorner,
                'left': win32c.LegendPosition.xlLegendPositionLeft,
                'right': win32c.LegendPosition.xlLegendPositionRight,
                'top': win32c.LegendPosition.xlLegendPositionTop
            }[position]
            chart.Legend.Position = position
        if seriesName:
            for index, elem in enumerate(seriesName):
                chart.SeriesCollection(index+1).Name = elem
        self.SaveWorkbook(output_wb_path)
        self.closeWorkBook()

    def SetChartHasLegend(self, src_wb_path: str, chartName: str, hasLegend: bool, output_wb_path: str = None) -> None:
        if output_wb_path == None:
            output_wb_path = src_wb_path
        self.OpenWorkbook(src_wb_path)
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        chart.HasLegend = hasLegend
        self.SaveWorkbook(output_wb_path)
        self.closeWorkBook()

    def SetChartType(self, src_wb_path: str, chartName: str, chartType: str, output_wb_path: str = None) -> None:
        # find the chart
        if output_wb_path == None:
            output_wb_path = src_wb_path
        self.OpenWorkbook(src_wb_path)
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        chart.ChartType = constants.ChartType[chartType]
        self.SaveWorkbook(output_wb_path)
        self.closeWorkBook()

    def SetChartSource(self, chartName: str, source: str) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        source = self.toRange(source)
        chart.SetSourceData(source)

    def SetChartBackgroundColor(self, src_wb_path :str, chartName: str, color: str, output_wb_path: str = None) -> None:
        # find the chart
        if output_wb_path == None:
            output_wb_path = src_wb_path
        self.OpenWorkbook(src_wb_path)
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        chart.ChartArea.Interior.ColorIndex = constants.ColorIndex[color]
        self.SaveWorkbook(output_wb_path)
        self.closeWorkBook()

    def ResizeChart(self, src_wb_path: str, chartName: str, width: float, height: float, output_wb_path: str = None) -> None:
        # find the chart
        if output_wb_path == None:
            output_wb_path = src_wb_path
        self.OpenWorkbook(src_wb_path)
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart.Width = width
        chart.Height = height
        self.SaveWorkbook(output_wb_path)
        self.closeWorkBook()

    def SetChartDataColor(self, src_wb_path : str, chartName: str, colorRGB: list, out_wb_path: str = None) -> None:
        # find the chart
        if output_wb_path == None:
            output_wb_path = src_wb_path
        self.OpenWorkbook(src_wb_path)
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        for series in chart.SeriesCollection():
            for point in series.Points():
                point.Format.Fill.ForeColor.RGB = win32api.RGB(*colorRGB)
        self.SaveWorkbook(output_wb_path)
        self.closeWorkBook()
    
    def HighlightDataPoints(self, src_wb_path: str, chartName: str, pointIndex: int, colorRGB: list, output_wb_path: str = None) -> None:
        # find the chart
        if output_wb_path == None:
            output_wb_path = src_wb_path
        self.OpenWorkbook(src_wb_path)
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        series = chart.SeriesCollection(1)
        point = series.Points(pointIndex)
        point.Format.Fill.ForeColor.RGB = win32api.RGB(*colorRGB)
        self.SaveWorkbook(output_wb_path)
        self.closeWorkBook()

    def SetDataSeriesType(self, src_wb_path: str,  chartName: str, seriesIndex: int, seriesType: str, output_wb_path: str = None) -> None:
        # find the chart
        if output_wb_path == None:
            output_wb_path = src_wb_path
        self.OpenWorkbook(src_wb_path)
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        series = chart.SeriesCollection(seriesIndex)
        series.ChartType = constants.ChartType[seriesType]
        self.SaveWorkbook(output_wb_path)
        self.closeWorkBook()

    def AddDataSeries(self, src_wb_path: str, src_sheet:str, chartName: str, xrange: str, yrange: str, output_wb_path: str = None) -> None:
        # find the chart
        if output_wb_path == None:
            output_wb_path = src_wb_path
        self.OpenWorkbook(src_wb_path) 
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        if src_sheet not in [sheet.Name for sheet in self.activeWB.Worksheets]:
            raise ValueError(f'Sheet {src_sheet} does not exist.')
        sheet = self.activeWB.Sheets(src_sheet)
        xrange = self.toRangeUsingSheet(xrange)
        yrange = self.toRangeUsingSheet(yrange)
        series = chart.SeriesCollection().NewSeries()
        series.XValues = xrange
        series.Values = yrange
        self.SaveWorkbook(output_wb_path)
        self.closeWorkBook()
        
    
    def RemoveDataSeries(self, src_wb_path: str, chartName: str, seriesIndex: int, output_wb_path: str = None) -> None:
        # find the chart
        if output_wb_path == None:
            output_wb_path = src_wb_path
        self.OpenWorkbook(src_wb_path)
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        series = chart.SeriesCollection(seriesIndex)
        series.Delete()
        self.SaveWorkbook(output_wb_path)
        self.closeWorkBook()

    def SetDataSeriesSource(self, src_wb_path: str, chartName: str, seriesIndex: int, xrange: str, yrange: str, output_wb_path: str = None) -> None:
        # find the chart
        if output_wb_path == None:
            output_wb_path = src_wb_path
        self.OpenWorkbook(src_wb_path)
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        xrange = self.toRangeUsingSheet(xrange)
        yrange = self.toRangeUsingSheet(yrange)
        series = chart.SeriesCollection(seriesIndex)
        series.XValues = xrange
        series.Values = yrange
        self.SaveWorkbook(output_wb_path)
        self.closeWorkBook()

    def AddChartErrorBars(self, src_wb_path: str, chartName: str, output_wb_path: str = None) -> None:
        # find the chart
        if output_wb_path == None:
            output_wb_path = src_wb_path
        self.OpenWorkbook(src_wb_path)
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        for series in chart.SeriesCollection():
            series.HasErrorBars = True
        self.SaveWorkbook(output_wb_path)
        self.closeWorkBook()

    def AddChartErrorBar(self, src_wb_path: str, chartName: str, seriesIndex: int, output_wb_path : str = None) -> None:
        # find the chart
        if output_wb_path == None:
            output_wb_path = src_wb_path
        self.OpenWorkbook(src_wb_path)
        if output_wb_path == None:
            output_wb_path = src_wb_path
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        series = chart.SeriesCollection(seriesIndex)
        series.HasErrorBars = True
        self.SaveWorkbook(output_wb_path)
        self.closeWorkBook()

    def RemoveChartErrorBars(self, src_wb_path: str, chartName: str, output_wb_path: str = None) -> None:
        # find the chart
        if output_wb_path == None:
            output_wb_path = src_wb_path
        self.OpenWorkbook(src_wb_path)
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        for series in chart.SeriesCollection():
            series.HasErrorBars = False
        self.SaveWorkbook(output_wb_path)
        self.closeWorkBook()

    def RemoveChartErrorBar(self, src_wb_path: str, chartName: str, seriesIndex: int, output_wb_path: str = None) -> None:
        # find the chart
        if output_wb_path == None:
            output_wb_path = src_wb_path
        self.OpenWorkbook(src_wb_path)
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        series = chart.SeriesCollection(seriesIndex)
        series.HasErrorBars = False
        self.SaveWorkbook(output_wb_path)
        self.closeWorkBook()

    def AddDataLabels(self, src_wb_path: str, chartName: str, output_wb_path: str = None) -> None:
        # find the chart
        if output_wb_path == None:
            output_wb_path = src_wb_path
        self.OpenWorkbook(src_wb_path)
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        for series in chart.SeriesCollection():
            series.HasDataLabels = True
        self.SaveWorkbook(output_wb_path)
        self.closeWorkBook()
    
    def RemoveDataLabels(self, src_wb_path: str, chartName: str, output_wb_path: str = None) -> None:
        # find the chart
        if output_wb_path == None:
            output_wb_path = src_wb_path
        self.OpenWorkbook(src_wb_path)
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        for series in chart.SeriesCollection():
            series.HasDataLabels = False
        self.SaveWorkbook(output_wb_path)
        self.closeWorkBook()

    def SetChartMarker(self,  src_wb_path: str, chartName: str, style: List[str] = None, size: Optional[float] = None, output_wb_path : str = None) -> None:
        if output_wb_path == None:
            output_wb_path = src_wb_path
        self.OpenWorkbook(src_wb_path)
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        styleIter = itertools.cycle(style)
        for series in chart.SeriesCollection():
            if style:
                series.MarkerStyle = constants.MarkerStyle[next(styleIter)]
            if size:
                series.MarkerSize = size
        self.SaveWorkbook(output_wb_path)
        self.closeWorkBook()

    def CopyPasteChart(self, src_wb_path: str, chartName, dest_sheet: str, destination: str, output_wb_path: str = None) -> None:
        # find the chart
        if output_wb_path == None:
            output_wb_path = src_wb_path
        self.OpenWorkbook(src_wb_path)
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        if dest_sheet not in [sheet.Name for sheet in self.activeWB.Worksheets]:
            raise ValueError(f'Sheet {dest_sheet} does not exist.')
        sheet = self.activeWB.Sheets(dest_sheet)
        chart = chart.Chart
        destination = self.toRange(sheet, destination)
        chart.ChartArea.Copy()
        destination.Select()
        destination.Parent.Paste()
        self.SaveWorkbook(output_wb_path)
        self.closeWorkBook()

    def GetBlankArea(self, sheetName: str):
        sheet = self.activeWB.Sheets(sheetName)
        chartRangeList = []
        for chart in sheet.ChartObjects():
            chartRangeList.append(sheet.Range(chart.TopLeftCell, chart.BottomRightCell))
        row, column = 1, 1
        checkScope = 5
        while True:
            cell1 = sheet.Cells(row,column)
            cell2 = sheet.Cells(row+checkScope,column+checkScope)
            checkRange = sheet.Range(cell1, cell2)
            if all(cell.Value is None for cell in checkRange) and all(self.activeAPP.Intersect(chartRange, checkRange) is None for chartRange in chartRangeList):
                break
            row += 1
            column += 1

        while row > 1:
            cell1 = sheet.Cells(row-1,column)
            cell2 = sheet.Cells(row-1,column+checkScope)
            checkRange = sheet.Range(cell1, cell2)
            if any(cell.Value is not None for cell in checkRange) or any(self.activeAPP.Intersect(chartRange, checkRange) is not None for chartRange in chartRangeList):
                break
            row -= 1
            
        while column > 1:
            cell1 = sheet.Cells(row,column-1)
            cell2 = sheet.Cells(row+checkScope,column-1)
            checkRange = sheet.Range(cell1, cell2)
            if any(cell.Value is not None for cell in checkRange) or any(self.activeAPP.Intersect(chartRange, checkRange) is not None for chartRange in chartRangeList):
                break
            column -= 1

        return sheet.Cells(row+1,column+1)

    
    def GetSheetsState(self) -> str:
        states = []
        for ws in self.activeWB.Worksheets:
            if ws.Range('A1').Value is None:
                cell_state = "Sheet \"{}\" {} has no content".format(ws.Name, '(active)' if ws.Name == self.activeWS.Name else '')
            else:
                NumberOfRows = ws.UsedRange.Rows.Count
                NumberOfColumns = ws.UsedRange.Columns.Count
                headers = ws.Range('A1', ws.Cells(1,NumberOfColumns)).Value
                if isinstance(headers, tuple):
                    headers = headers[0]
                else:
                    headers = [headers]
                headers = {get_column_letter(i): header for i, header in enumerate(headers, start=1)}
                
                cell_state = 'Sheet \"{}\" has {} columns (Headers are {}) and {} rows (1 header row and {} data rows)'.format(ws.Name, NumberOfColumns, ', '.join(["{}: \"{}\"".format(col_letter, header) for col_letter, header in headers.items()]), NumberOfRows, NumberOfRows-1)
            
            # Add chart descriptions
            # Iterate through the shapes to find chart objects
            chart_names = []
            for shape in ws.Shapes:
                if shape.HasChart:
                    chart = shape.Chart
                    chart_name = chart.Name
                    chart_names.append(chart_name[chart_name.find(' ')+1:])
                    
            chartNameString = ' and this sheet has the charts whose names are "{}"'.format('", "'.join(chart_names)) if len(chart_names) > 0 else ''
            
            # Iterate through the pivot tables and print their names
            pt_names = []
            for pivot_table in ws.PivotTables():
                pt_names.append(pivot_table.Name)

            if len(pt_names) > 0:
                ptNameString = ' the pivot tables whose names are "{}"'.format('", "'.join(pt_names))
                if len(chart_names) == 0:
                    ptNameString = ' and this sheet has' + ptNameString
                else:
                    ptNameString = ' and' + ptNameString
            else:
                ptNameString = ''
            
            states.append("{}{}{}.".format(cell_state, chartNameString, ptNameString))
                                                                                                                       
        return "Sheet state: " + ' '.join(states)



# app = Charts_App()
# app.OpenWorkbook("./IncomeStatement.xlsx")
# app.CreateChart("./IncomeStatement.xlsx", "Sheet1!B2:C9", "Sheet1", "Barchart", "profits", 2, [3],"./IncomeStatement.xlsx")