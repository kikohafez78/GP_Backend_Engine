
#=================================
#openpyxl lib
# from openpyxl import load_workbook
# from openpyxl.worksheet.datavalidation import DataValidation
# from openpyxl.worksheet.page import PrintPageSetup,PageMargins
# from openpyxl.utils import get_column_letter,range_boundaries
# from openpyxl.utils.cell import coordinate_from_string
# from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle, Protection
# from openpyxl.formatting.rule import CellIsRule
# from openpyxl.drawing.image import Image
#=================================

# class formatting(object):
#     """
# Format cells: Change the appearance of cells, such as font, color, alignment, etc.
# Set data type: Specify the data type for a range of cells (e.g., text, number, date).
# Delete format: Remove formatting from cells.
# Change page layout: Adjust page settings such as orientation, margins, and print area.
# Set border: Add borders to cells or ranges of cells.
# Resize cells: Adjust the width and height of cells.
# Conditional formatting: Apply formatting to cells based on specific conditions or criteria.
# Lock and unlock: Protect cells from editing or unlock them for editing.
# Protect: Protect the workbook or specific sheets from unauthorized changes.
# Unprotect: Remove protection from the workbook or specific sheets.
# Drop-down list: Create a drop-down list control in a cell.
# Data validation: Specify restrictions or rules for the data entered into cells.
# Display formulas: Show or hide formulas in cells.
# Wrap text: Wrap text within cells to display long content.
# Unwrap text: Remove text wrapping from cells.
# Autofit: Automatically adjust the width of columns or height of rows to fit the content.
#     """
#     def __init__(self):
#         super().__init__()
#         return
    
#     @staticmethod
#     def format_cells(workbook_path,sheet_name, data_range, bold: bool = True,  italic: bool = True, color: str = "FF0000", horizontalAlignment: str = "center", verticalAlignment: str = "center", fillColor: str = "FFFF00", fillType: str = "solid", borderStyle: str = "thin", output_workbook_path=None):
#         if output_workbook_path == None:
#             output_workbook_path = workbook_path
#         workbook = load_workbook(workbook_path)
#         sheet = workbook[sheet_name]
#         start_row_p1, start_column_p1, end_col_p1, end_row_p1 = range_boundaries(data_range)
#         for row_idx in range(start_row_p1, end_row_p1 + 1):
#             for col_idx in range(start_column_p1, end_col_p1 + 1):
#                 cell = sheet.cell(row = row_idx, column = col_idx)
#                 cell.font = Font(bold = bold, italic = italic, color = color)  # Bold, italic, red font color
#                 cell.alignment = Alignment(horizontal = horizontalAlignment, vertical = verticalAlignment)  # Center align content
#                 cell.border = Border(left = Side(style = borderStyle), right = Side(style = borderStyle), top = Side(style = borderStyle), bottom=Side(style = borderStyle))  # Thin border around cell
#                 cell.fill = PatternFill(start_color = fillColor, end_color = fillColor, fill_type = fillType)  # Yellow background color
#         workbook.save(output_workbook_path)
#         workbook.close()
        
#     @staticmethod
#     def set_data_type(workbook_path: str, sheet_name: str, range_start, range_end, datatype, output_workbook_path = None):
#         if output_workbook_path == None:
#             output_workbook_path = workbook_path
#         workbook = load_workbook(workbook_path)
#         sheet = workbook[sheet_name]
#         start_row, start_column = range_boundaries(range_start)
#         end_row, end_column = range_boundaries(range_end)
#         for row in range(start_row, end_row + 1):
#             for column in range(start_column, end_column + 1):
#                 cell = sheet.cell(row=row, column=column)
#                 if column == start_column + 1:  # Check if column is second column in the range
#                     cell.number_format = datatype  # Set data type as integer
#                 else:
#                     cell.number_format = "@"  # Set data type as text
#         workbook.save(output_workbook_path)
#         workbook.close()
        
#     @staticmethod
#     def remove_format(workbook_path, sheet_name, range_start, range_end, output_workbook_path = None):
#         if output_workbook_path == None:
#             output_workbook_path = workbook_path        
#         workbook = load_workbook(workbook_path)
#         sheet = workbook[sheet_name]
#         start_row, start_column = range_boundaries(range_start)
#         end_row, end_column = range_boundaries(range_end)
#         for row in range(start_row, end_row + 1):
#             for column in range(start_column, end_column + 1):
#                 cell = sheet.cell(row=row, column=column)
#                 cell.alignment = NamedStyle(name = 'Normal').alignment  # Reset alignment
#                 cell.font = NamedStyle(name = 'Normal').font  # Reset font
#                 cell.fill = NamedStyle(name = 'Normal').fill  # Reset fill
#                 cell.border = NamedStyle(name = 'Normal').border  # Reset border
#                 cell.number_format = "General"  # Reset number format
#         workbook.save(output_workbook_path)
#         workbook.close()
        
#     @staticmethod
#     def change_page_layout(workbook_path, sheet_name, orientation='portrait', paper_size='A4', margin_top=0.5, margin_right=0.5, margin_bottom=0.5, margin_left=0.5, print_area=None, output_workbook_path= None):
#         if output_workbook_path == None:
#             output_workbook_path = workbook_path        
#         workbook = load_workbook(workbook_path)
#         sheet = workbook[sheet_name]
#         page_setup = PrintPageSetup(sheet)
#         page_margins = PageMargins()
#         page_setup.orientation = orientation
#         page_setup.paperSize = paper_size
#         page_margins.top = margin_top
#         page_margins.right = margin_right
#         page_margins.bottom = margin_bottom
#         page_margins.left = margin_left
#         if print_area:
#             sheet.print_area = print_area
#         workbook.save(output_workbook_path)
#         workbook.close()
        
#     @staticmethod
#     def set_border(workbook_path, sheet_name, range_start, range_end, style: str, color: str, output_workbook_path = None):
#         if output_workbook_path == None:
#             output_workbook_path = workbook_path        
#         workbook = load_workbook(workbook_path)
#         sheet = workbook[sheet_name]
#         start_row, start_column = coordinate_from_string(range_start)
#         end_row, end_column = coordinate_from_string(range_end)
#         thin_border = Border(left = Side(style = style, color = color), 
#                              right = Side(style = style, color = color), 
#                              top = Side(style = style, color = color), 
#                              bottom = Side(style = style, color = color))
#         for row in sheet.iter_rows(min_row=start_row, min_col=start_column, max_row=end_row, max_col=end_column):
#             for cell in row:
#                 cell.border = thin_border
#         workbook.save(output_workbook_path)
#         workbook.close()
        
#     @staticmethod
#     def resize_cells(workbook_path, sheet_name, start_cell, end_cell, column_width=None, row_height=None, output_workbook_path= None):
#         if output_workbook_path == None:
#             output_workbook_path = workbook_path        
#         workbook = load_workbook(workbook_path)
#         sheet = workbook[sheet_name]
#         start_row, start_column = coordinate_from_string(start_cell)
#         end_row, end_column = coordinate_from_string(end_cell)
#         if column_width is not None:
#             for col in range(start_column, end_column + 1):
#                 sheet.column_dimensions[get_column_letter(col)].width = column_width
#         if row_height is not None:
#             for row in range(start_row, end_row + 1):
#                 sheet.row_dimensions[row].height = row_height
#         workbook.save(output_workbook_path)
#         workbook.close()
        
#     @staticmethod
#     def apply_conditional_formatting(workbook_path, sheet_name, range_string, condition, style, output_workbook_path = None):
#         if output_workbook_path == None:
#             output_workbook_path = workbook_path        
#         workbook = load_workbook(workbook_path)
#         sheet = workbook[sheet_name]
#         rule = CellIsRule(operator = condition.operator, formula = condition.formula, stopIfTrue = False, font = style["font"],fill=style["fill"],border=style["border"])
#         sheet.conditional_formatting.add(range_string, rule)
#         workbook.save(output_workbook_path)
#         workbook.close()
        
#     @staticmethod
#     def lock_cells(workbook_path, sheet_name, range_string, lock, output_workbook_path = None):
#         if output_workbook_path == None:
#             output_workbook_path = workbook_path        
#         workbook = load_workbook(workbook_path)
#         sheet = workbook[sheet_name]
#         for row in sheet[range_string]:
#                 row.protection = Protection(locked = lock)
#         sheet.protection.sheet = True
#         workbook.save(output_workbook_path)
#         workbook.close()
    
#     @staticmethod
#     def unlock_cells(workbook_path, sheet_name, range_string, output_workbook_path = None):
#         if output_workbook_path == None:
#             output_workbook_path = workbook_path        
#         workbook = load_workbook(workbook_path)
#         sheet = workbook[sheet_name]
#         for row in sheet[range_string]:
#                 row.protection = None
#         sheet.protection.sheet = False
#         workbook.save(output_workbook_path)
#         workbook.close()

#     @staticmethod
#     def protect(workbook_path, sheet_name, password=None,output_workbook_path = None):
#         if output_workbook_path == None:
#             output_workbook_path = workbook_path        
#         workbook = load_workbook(workbook_path)
#         sheet = workbook[sheet_name]
#         if password:
#             sheet.protection.sheet = True
#             sheet.protection.password = password
#         else:
#             sheet.protection.sheet = True
#         workbook.save(output_workbook_path) 
#         workbook.close()   
        
#     @staticmethod
#     def unprotect(workbook_path, sheet_name, output_workbook_path = None):
#         if output_workbook_path == None:
#             output_workbook_path = workbook_path        
#         workbook = load_workbook(workbook_path)
#         sheet = workbook[sheet_name]
#         sheet.protection.sheet = None
#         sheet.protection.password = None
#         workbook.save(output_workbook_path)
#         workbook.close()
        
#     @staticmethod
#     def dropdown_list(workbook_path, sheet_name, start_cell, end_cell, values, output_workbook_path = None):
#         if output_workbook_path == None:
#             output_workbook_path = workbook_path        
#         workbook = load_workbook(workbook_path)
#         sheet = workbook[sheet_name]
#         data_validation = DataValidation(type="list", formula1='"' + ','.join(values) + '"', showDropDown=True)
#         for row in sheet[start_cell:end_cell]:
#             for cell in row:
#                 sheet.add_data_validation(data_validation)
#                 data_validation.add(cell)
#         workbook.save(output_workbook_path)
#         workbook.close()
        
#     @staticmethod
#     def set_border(workbook_path, sheet_name, range_start, range_end, wrap_text: bool = True,output_workbook_path = None):
#         if output_workbook_path == None:
#             output_workbook_path = workbook_path        
#         workbook = load_workbook(workbook_path)
#         sheet = workbook[sheet_name]
#         start_row, start_column = coordinate_from_string(range_start)
#         end_row, end_column = coordinate_from_string(range_end)
#         for row in sheet.iter_rows(min_row=start_row, min_col=start_column, max_row=end_row, max_col=end_column):
#             for cell in row:
#                 cell.alignment.wrap_text = wrap_text
#         workbook.save(output_workbook_path) 
#         workbook.close()
    
#     @staticmethod
#     def data_validation(workbook_path, sheet_name, range_start, range_end, validation_type, criteria, prompt=None, error_message=None,output_workbook_path = None):
#         if output_workbook_path == None:
#             output_workbook_path = workbook_path        
#         workbook = load_workbook(workbook_path)
#         sheet = workbook[sheet_name]
#         data_validation = DataValidation(type=validation_type, formula1=criteria, promptTitle=prompt, prompt=prompt, errorTitle=error_message, error=error_message)
#         sheet.add_data_validation(data_validation)
#         data_validation.add(sheet[range_start:range_end])
#         workbook.save(output_workbook_path)
#         workbook.close()
        
        
#     @staticmethod
#     def display_formulas(workbook_path, sheet_name, output_workbook_path, display: bool = True):
#         if output_workbook_path == None:
#             output_workbook_path = workbook_path         
#         workbook = load_workbook(workbook_path)
#         sheet = workbook[sheet_name]
#         sheet.sheet_view.showFormulas = display
#         workbook.save(output_workbook_path)
#         workbook.close()
        
#     @staticmethod #revise later
#     def wrap_text(workbook_path, sheet_name, range_start, range_end, wrapper,output_workbook_path):
#         if output_workbook_path == None:
#             output_workbook_path = workbook_path         
#         workbook = load_workbook(workbook_path)
#         sheet = workbook[sheet_name]
#         for row in sheet[range_start:range_end]:
#             row.alignment = Alignment(wrapText = wrapper)
#         workbook.save(output_workbook_path)
#         workbook.close()
        
#     @staticmethod
#     def add_data_validation(workbook_path, sheet_name, range_start, range_end, validation_type, criteria,output_workbook_path, prompt=None, error_message=None):
#         if output_workbook_path == None:
#             output_workbook_path = workbook_path         
#         workbook = load_workbook(workbook_path)
#         sheet = workbook[sheet_name]
#         data_validation = DataValidation(type=validation_type, formula1=criteria, promptTitle=prompt, prompt=prompt, errorTitle=error_message, error=error_message)
#         sheet.add_data_validation(data_validation)
#         data_validation.add(sheet[range_start:range_end])
#         workbook.save(output_workbook_path)
#         workbook.close()
    
#     def autofit_columns(workbook_path, sheet_name, output_workbook_path):
#         if output_workbook_path == None:
#             output_workbook_path = workbook_path         
#         workbook = load_workbook(workbook_path)
#         sheet = workbook[sheet_name]
#         for column in sheet.columns:
#             max_length = 0
#             column_letter = get_column_letter(column[0].column)
#             for cell in column:
#                 try:
#                     if len(str(cell.value)) > max_length:
#                         max_length = len(cell.value)
#                 except:
#                     pass
#             adjusted_width = (max_length + 2) * 1.2 
#             sheet.column_dimensions[column_letter].width = adjusted_width
#         workbook.save(output_workbook_path)
#         workbook.close()
        



from xlwings import constants as win32c
from constants import constants
import win32com.client as win32
from typing import List
from openpyxl.utils import get_column_letter
import os
import sys
script_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.append(script_dir)
from typing import Optional

class formatting_App():
    def __init__(self, app = 'excel', api_doc = None) -> None:
        super().__init__()
        self.appName = app
        self.__excel = None
        self.__currentWB = None

        if api_doc is not None:
            for key, value in api_doc.items():
                if value.get('display') is not None:
                    setattr(self, value['display'], self.__getattribute__(key))

    @property
    def activeAPP(self):
        if not self.__excel:
            try:
                self.__excel = win32.Dispatch('Excel.Application') if self.appName == 'excel' else win32.Dispatch('ket.Application')
                self.__excel.DisplayAlerts = False
                self.__excel.Visible = False
            except:
                raise Exception('{} is not running.'.format(self.appName))
        return self.__excel
    
    @property
    def activeWB(self):
        if self.__currentWB is not None:
            return self.__currentWB
        return self.activeAPP.ActiveWorkbook
    
    @activeWB.setter
    def activeWB(self, wb):
        self.__currentWB = wb
    
    @property
    def activeWS(self):
        return self.activeWB.ActiveSheet
    
    def toRangeUsingSheet(self, source: str):
        if '!' in source:
            sheet_name, source = source.split('!')
            sheet_name = sheet_name.strip("'") 
            if sheet_name not in [sheet.Name for sheet in self.activeWB.Worksheets]:
                raise ValueError(f'Sheet {sheet_name} does not exist.')
            sheet = self.activeWB.Sheets(sheet_name)
        else:
            raise Exception('The range must contain a sheet name.')
        
        if source.isdigit():
            return sheet.Rows(source)
        elif source.isalpha():
            return sheet.Columns(source)
        elif ':' in source or source.isalnum():
            return sheet.Range(source)
        
    def toRange(self, sheet,  source: str):
        if source.isdigit():
            return sheet.Rows(source)
        elif source.isalpha():
            return sheet.Columns(source)
        elif ':' in source or source.isalnum():
            return sheet.Range(source)
    
    def OpenWorkbook(self, path: str) -> None:
        self.__currentWB = self.activeAPP.Workbooks.Open(os.path.abspath(path))

    def SaveWorkbook(self, path: str) -> None:
        self.activeWB.SaveAs(os.path.abspath(path))
    
    def closeWorkBook(self) -> None:
        self.activeWB.Close()
     
        
    def format_cells(self,workbook_path, sheet_name, cell_range: str, font: Optional[str] = None, fontSize: Optional[float] = None,
                    color: Optional[int] = None, fillColor: Optional[int] = None, bold: Optional[bool] = None,
                    italic: Optional[bool] = None, underline: Optional[bool] = None, horizontalAlignment: Optional[str] = None, output_workbook_path: str = "") -> None:
        # self.OpenWorkbook(workbook_path)
        sheet = self.activeWB.Sheets(sheet_name)
        source = self.toRange(sheet, cell_range)
        if font:
            source.Font.Name = font
        if fontSize:
            source.Font.Size = fontSize
        if color:
            source.Font.ColorIndex = constants.ColorIndex[color]
        if fillColor:
            source.Interior.ColorIndex = constants.ColorIndex[fillColor]
        if not bold is None:
            source.Font.Bold = bold
        if not italic is None:
            source.Font.Italic = italic
        if not underline is None:
            source.Font.Underline = win32c.UnderlineStyle.xlUnderlineStyleSingle if underline else win32c.UnderlineStyle.xlUnderlineStyleNone
        if horizontalAlignment:
            source.HorizontalAlignment = constants.HorizontalAlignment[horizontalAlignment]
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
        return f"formatted cells in the range {cell_range} for sheet {sheet_name}"
    
        
    def delete_format(self,workbook_path, sheet_name, cell_range: str , output_workbook_path: str = "") -> None:
        # self.OpenWorkbook(workbook_path)
        sheet = self.activeWB.Sheets(sheet_name)
        source = self.toRange(sheet, cell_range)
        source.ClearFormats() 
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
        return f"deleted format of cells in the range {cell_range} for sheet {sheet_name}"
    
    def set_data_type(self,workbook_path, sheet_name, dataType: str, cell_range: str , output_workbook_path: str = "") -> None:
        # self.OpenWorkbook(workbook_path)
        sheet = self.activeWB.Sheets(sheet_name)
        source = self.toRange(sheet, cell_range)
        source.NumberFormat = constants.DataType[dataType]
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
        return f"dtatype format of cells in the range {cell_range} is changed to {dataType} for sheet {sheet_name}"
    
    def change_page_layout(self,workbook_path, sheet_name,paper_size, orientation , output_workbook_path: str = "") -> None:
        # self.OpenWorkbook(workbook_path)
        sheet = self.activeWB.Sheets(sheet_name)
        sheet.PageSetup.Orientation = constants.PageOrientation[orientation]
        sheet.PageSetup.PaperSize = constants.PaperSize[paper_size]
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
        return f"page layout for sheet '{sheet_name}' is changed for orientation {orientation} and paper size {paper_size}"
    
    def set_border(self, workbook_path, sheet_name, cell_range, color: str, weight: str , output_workbook_path = None):
        # self.OpenWorkbook(workbook_path)
        sheet = self.activeWB.Sheets(sheet_name)
        source = self.toRange(sheet, cell_range)
        source = self.toRange(source)
        source.BorderAround(ColorIndex=constants.ColorIndex[color], Weight=constants.BorderWeight[weight])
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
        return f"border for range {cell_range} in sheet '{sheet_name}' is set to {weight} weight and {color} color"
    
    def data_validation(self, workbook_path, sheet_name, cell_range, type: str, formula1: str, output_workbook_path = None) -> None:
        # self.OpenWorkbook(workbook_path)
        sheet = self.activeWB.Sheets(sheet_name)
        source = self.toRange(sheet, cell_range)
        handle = source.Validation.Add(constants.ValidationType[type], Formula1 = formula1)
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
        return f"data validation for range {cell_range} in sheet '{sheet_name}' is set to {type} type and {formula1} formula"
    
    def display_formula(self, workbook_path, display: bool, output_workbook_path):
        # self.OpenWorkbook(workbook_path)
        self.activeAPP.ActiveWindow.DisplayFormulas = display
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
        return f"formula display for workbook is set to {'visible' if display else 'invisible'}"
    
    def wrap_unwrap_text(self, workbook_path, sheet_name, cell_range, output_workbook_path = None, wrap = True) -> None:
        # self.OpenWorkbook(workbook_path)
        sheet = self.activeWB.Sheets(sheet_name)
        source = self.toRange(sheet, cell_range)
        source.WrapText = wrap
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
        return f"data validation for range {cell_range} in sheet {sheet_name} is set to {'wrapped' if wrap else 'unwrapped'}"
    
    def autofit(self, workbook_path, sheet_name, cell_range, output_workbook_path = None) -> None:
        # self.OpenWorkbook(workbook_path)
        sheet = self.activeWB.Sheets(sheet_name)
        source = self.toRange(sheet, cell_range)
        source.AutoFit()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
        return f"the range {cell_range} in sheet {sheet_name} is Autofitted"
    
    def resize_cells(self, workbook_path, sheet_name, cell_range, width: Optional[int] = None, height: Optional[int] = None, output_workbook_path = None) -> None:
        # self.OpenWorkbook(workbook_path)
        sheet = self.activeWB.Sheets(sheet_name)
        source = self.toRange(sheet, cell_range)
        if height:
            source.RowHeight = height
        if width:
            source.ColumnWidth = width
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
        return f"the range {cell_range} in sheet {sheet_name} is Autofitted"
    
    def conditional_formatting(self, workbook_path, sheet_name, cell_range,formula: str,
                            bold: Optional[bool] = None, color: Optional[str] = None,
                            fillColor: Optional[str] = None, italic: Optional[bool] = None, underline: Optional[bool] = None, output_workbook_path = None) -> None:
        # self.OpenWorkbook(workbook_path)
        sheet = self.activeWB.Sheets(sheet_name)
        source = self.toRange(sheet, cell_range)
        handle = source.FormatConditions.Add(Type = constants.FormatConditionType['expression'], Formula1 = formula)
        if color:
            handle.Font.ColorIndex = constants.ColorIndex[color.lower()]
        if fillColor:
            handle.Interior.ColorIndex = constants.ColorIndex[fillColor.lower()]
        if not bold is None:
            handle.Font.Bold = bold
        if not italic is None:
            handle.Font.Italic = italic
        if not underline is None:
            handle.Font.Underline = win32c.UnderlineStyle.xlUnderlineStyleSingle if underline else win32c.UnderlineStyle.xlUnderlineStyleNone
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
        return f"conditional formatting is applied for range {cell_range} in sheet {sheet_name}" #<== continue doc
    
    def lock_unlock_cells(self, workbook_path, sheet_name, cell_range, lock, output_workbook_path = None) -> None:
        # self.OpenWorkbook(workbook_path)
        sheet = self.activeWB.Sheets(sheet_name)
        source = self.toRange(sheet, cell_range)
        source.Locked = lock
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
        return f"the range {cell_range} in sheet {sheet_name} is {'locked' if lock else 'unlocked'}"
    
    
    def protect_unprotect_cells(self, workbook_path, sheet_name, cell_range, protect, password, output_workbook_path = None) -> None:
        # self.OpenWorkbook(workbook_path)
        sheet = self.activeWB.Sheets(sheet_name)
        if protect:
            sheet.Protect(password)
        else:
            sheet.Unprotect(password)
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
        return f"the {sheet_name} sheet is {f'protected with {password}' if protect else 'unprotected'}"
    
    
    def dropdown_list(self, workbook_path, sheet_name, cell_range, dropdown_values, output_workbook_path = None):
        # self.OpenWorkbook(workbook_path)
        sheet = self.activeWB.Sheets(sheet_name)
        dropdown_range = self.toRange(sheet, cell_range)
        try:
            values_str = ','.join(dropdown_values)
            dropdown_range.Validation.Delete()
            dropdown_range.Validation.Add(3, 1, 1, values_str)
            dropdown_range.Validation.IgnoreBlank = True
            dropdown_range.Validation.InCellDropdown = True
        except Exception as e:
            print(f"Error: {e}")
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
        return f"" #<== finish this
    
    
    
    
    
    
    
        