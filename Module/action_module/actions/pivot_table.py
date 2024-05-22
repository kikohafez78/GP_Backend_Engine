from xlwings import constants as win32c
from constants import constants
import win32com.client as win32
from typing import List
from openpyxl.utils import get_column_letter
import os
import sys
script_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.append(script_dir)

class Pivot_App():
    def __init__(self, app = 'excel', api_doc = None) -> None:
        super().__init__()
        self.appName = app
        self.__excel = None
        self.__currentWB = None

        if api_doc is not None:
            for key, value in api_doc.items():
                if value.get('display') is not None:
                    setattr(self, value['display'], self.__getattribute__(key))

    @property
    def activeAPP(self):
        if not self.__excel:
            try:
                self.__excel = win32.Dispatch('Excel.Application') if self.appName == 'excel' else win32.Dispatch('ket.Application')
                self.__excel.DisplayAlerts = False
                self.__excel.Visible = False
            except:
                raise Exception('{} is not running.'.format(self.appName))
        return self.__excel
    
    @property
    def activeWB(self):
        if self.__currentWB is not None:
            return self.__currentWB
        return self.activeAPP.ActiveWorkbook
    
    @activeWB.setter
    def activeWB(self, wb):
        self.__currentWB = wb
    
    @property
    def activeWS(self):
        return self.activeWB.ActiveSheet
    
    def toRangeUsingSheet(self, source: str):
        if '!' in source:
            sheet_name, source = source.split('!')
            sheet_name = sheet_name.strip("'") 
            if sheet_name not in [sheet.Name for sheet in self.activeWB.Worksheets]:
                raise ValueError(f'Sheet {sheet_name} does not exist.')
            sheet = self.activeWB.Sheets(sheet_name)
        else:
            raise Exception('The range must contain a sheet name.')
        
        if source.isdigit():
            return sheet.Rows(source)
        elif source.isalpha():
            return sheet.Columns(source)
        elif ':' in source or source.isalnum():
            return sheet.Range(source)
        
    def toRange(self, sheet,  source: str):
        if source.isdigit():
            return sheet.Rows(source)
        elif source.isalpha():
            return sheet.Columns(source)
        elif ':' in source or source.isalnum():
            return sheet.Range(source)
    
    def OpenWorkbook(self, path: str) -> None:
        self.__currentWB = self.activeAPP.Workbooks.Open(os.path.abspath(path))

    def SaveWorkbook(self, path: str) -> None:
        self.activeWB.SaveAs(os.path.abspath(path))
    
    def closeWorkBook(self) -> None:
        self.activeWB.Close()

    def CreatePivotTable(self, src_wb_path: str, src_wb_sheet: str, source: str, destSheet: str, name: str,
                        RowField: List = [], ColumnField: List = [],
                        PageField: List = [], DataField: List = [],
                        summarizeFunction = 'sum', output_wb_path: str = None) -> None:
        # check if the pivot table name exists
        if output_wb_path == None:
            output_wb_path = src_wb_path
        self.OpenWorkbook(src_wb_path)
        for sheet in self.activeWB.Worksheets:
            for pt in sheet.PivotTables():
                if pt.Name == name:
                    raise ValueError(f'Pivot table {name} already exists. Please choose a different name.')
        # check if the destSheet exists
        if destSheet not in [sheet.Name for sheet in self.activeWB.Worksheets]:
            raise ValueError(f'Sheet {destSheet} does not exist.')
        if src_wb_sheet not in [sheet.Name for sheet in self.activeWB.Worksheets]:
            raise ValueError(f'Sheet {src_wb_sheet} does not exist.')
        sheet = self.activeWB.Worksheets(destSheet)
        sourceRange = self.toRange(sheet, source)
        sourceRange_sheet = sourceRange.Worksheet
        # Sometimes the LLM misses the header row, so we manually add it
        if sourceRange.Row != 1:
            new_starting_cell = sourceRange_sheet.Cells(1, sourceRange.Column)
            sourceRange = sourceRange_sheet.Range(new_starting_cell, sourceRange_sheet.Cells(new_starting_cell.Row + sourceRange.Rows.Count, sourceRange.Column + sourceRange.Columns.Count - 1))
            
        pc = self.activeWB.PivotCaches().Create(SourceType=win32c.PivotTableSourceType.xlDatabase, SourceData=sourceRange)
        destRange = self.GetBlankArea(destSheet)

        pt = pc.CreatePivotTable(TableDestination=destRange, TableName=name)
        for field in RowField:
            try:
                pt.PivotFields(field).Orientation = win32c.PivotFieldOrientation.xlRowField
            except:
                raise Exception(f"The field {field} is not included in the pivot table source range and cannot be selected as a row field!")
        for field in ColumnField:
            try:
                pt.PivotFields(field).Orientation = win32c.PivotFieldOrientation.xlColumnField
            except:
                raise Exception(f"The field {field} is not included in the pivot table source range and cannot be selected as a column field!")

        for field in PageField:
            try:
                pt.PivotFields(field).Orientation = win32c.PivotFieldOrientation.xlPageField
            except:
                raise Exception(f"The field {field} is not included in the pivot table source range and cannot be selected as a page field!")
        
        for field in DataField:
            try:
                pt.AddDataField(pt.PivotFields(field), Function = constants.SummarizationFunction[summarizeFunction])
            except:
                raise Exception(f"The field {field} is not included in the pivot table source range and cannot be selected as a data field!")
        self.SaveWorkbook(output_wb_path)
        self.closeWorkBook()
        return f"created a pivot table in sheet {destSheet} with the following properties" #<==== finsih this
         
    def GetBlankArea(self, sheetName: str):
        sheet = self.activeWB.Sheets(sheetName)
        chartRangeList = []
        for chart in sheet.ChartObjects():
            chartRangeList.append(sheet.Range(chart.TopLeftCell, chart.BottomRightCell))
        row, column = 1, 1
        checkScope = 5
        while True:
            cell1 = sheet.Cells(row,column)
            cell2 = sheet.Cells(row+checkScope,column+checkScope)
            checkRange = sheet.Range(cell1, cell2)
            if all(cell.Value is None for cell in checkRange) and all(self.activeAPP.Intersect(chartRange, checkRange) is None for chartRange in chartRangeList):
                break
            row += 1
            column += 1

        while row > 1:
            cell1 = sheet.Cells(row-1,column)
            cell2 = sheet.Cells(row-1,column+checkScope)
            checkRange = sheet.Range(cell1, cell2)
            if any(cell.Value is not None for cell in checkRange) or any(self.activeAPP.Intersect(chartRange, checkRange) is not None for chartRange in chartRangeList):
                break
            row -= 1
            
        while column > 1:
            cell1 = sheet.Cells(row,column-1)
            cell2 = sheet.Cells(row+checkScope,column-1)
            checkRange = sheet.Range(cell1, cell2)
            if any(cell.Value is not None for cell in checkRange) or any(self.activeAPP.Intersect(chartRange, checkRange) is not None for chartRange in chartRangeList):
                break
            column -= 1

        return sheet.Cells(row+1,column+1)

        
    def remove_pivot_table(self, src_wb_path: str, name: str, output_wb_path: str = None) -> None:
        # find the pivot table
        if output_wb_path == None:
            output_wb_path = src_wb_path
        self.OpenWorkbook(src_wb_path)
        pt = None
        for sheet in self.activeWB.Worksheets:
            for pt in sheet.PivotTables():
                if pt.Name == name:
                    break
        if pt is None or pt.Name != name:
            raise ValueError(f'Pivot table {name} does not exist.')
        pt.TableRange2.Clear()
        self.SaveWorkbook(output_wb_path)
        self.closeWorkBook()
        return f"removed a pivot table with name {name}" #<==== finsih this

    def set_summary_type(self, src_wb_path: str, name: str, field: str, func: str, output_wb_path: str = None) -> None:
        if output_wb_path == None:
            output_wb_path = src_wb_path
        self.OpenWorkbook(src_wb_path)
        pt = self.activeWS.PivotTables(name)
        pt.PivotFields(field).Function = constants.ConsolidationFunction[func]
        self.SaveWorkbook(output_wb_path)
        self.closeWorkBook()
        return f"changed pivot table with name {name} summary type to {func} summary" #<==== finsih this
        
    def sort_pivot_table(self, src_wb_path: str, name: str, field: str, key: str, order: str = 'ascending', output_wb_path: str = None) -> None:
        if output_wb_path == None:
            output_wb_path = src_wb_path
        self.OpenWorkbook(src_wb_path)
        pt = self.activeWS.PivotTables(name)
        order = constants.SortOrder[order]
        pt.PivotFields(field).AutoSort(order, key)
        self.SaveWorkbook(output_wb_path)
        self.closeWorkBook()
        return f"sorted pivot table with name {name} with the following feild {field} in {order} order and with key {key}" #<==== finsih this

    def GetSheetsState(self) -> str:
        states = []
        for ws in self.activeWB.Worksheets:
            if ws.Range('A1').Value is None:
                cell_state = "Sheet \"{}\" {} has no content".format(ws.Name, '(active)' if ws.Name == self.activeWS.Name else '')
            else:
                NumberOfRows = ws.UsedRange.Rows.Count
                NumberOfColumns = ws.UsedRange.Columns.Count
                headers = ws.Range('A1', ws.Cells(1,NumberOfColumns)).Value
                if isinstance(headers, tuple):
                    headers = headers[0]
                else:
                    headers = [headers]
                headers = {get_column_letter(i): header for i, header in enumerate(headers, start=1)}
                
                cell_state = 'Sheet \"{}\" has {} columns (Headers are {}) and {} rows (1 header row and {} data rows)'.format(ws.Name, NumberOfColumns, ', '.join(["{}: \"{}\"".format(col_letter, header) for col_letter, header in headers.items()]), NumberOfRows, NumberOfRows-1)
            
            # Add chart descriptions
            # Iterate through the shapes to find chart objects
            chart_names = []
            for shape in ws.Shapes:
                if shape.HasChart:
                    chart = shape.Chart
                    chart_name = chart.Name
                    chart_names.append(chart_name[chart_name.find(' ')+1:])
                    
            chartNameString = ' and this sheet has the charts whose names are "{}"'.format('", "'.join(chart_names)) if len(chart_names) > 0 else ''
            
            # Iterate through the pivot tables and print their names
            pt_names = []
            for pivot_table in ws.PivotTables():
                pt_names.append(pivot_table.Name)

            if len(pt_names) > 0:
                ptNameString = ' the pivot tables whose names are "{}"'.format('", "'.join(pt_names))
                if len(chart_names) == 0:
                    ptNameString = ' and this sheet has' + ptNameString
                else:
                    ptNameString = ' and' + ptNameString
            else:
                ptNameString = ''
            
            states.append("{}{}{}.".format(cell_state, chartNameString, ptNameString))
                                                                                                                       
        return "Sheet state: " + ' '.join(states)

