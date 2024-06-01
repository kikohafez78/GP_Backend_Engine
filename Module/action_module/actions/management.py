# import openpyxl as excel
# from openpyxl.utils import  range_boundaries, column_index_from_string
# from openpyxl.utils.cell import coordinate_from_string,get_column_letter
# import xlsxwriter as xw
# import pandas as pd
# import spire.xls as xls
# import re
# from openpyxl.workbook.defined_name import DefinedName
# from openpyxl.utils import quote_sheetname, absolute_coordinate
# from openpyxl.workbook.defined_name import DefinedName
# from openpyxl.utils import quote_sheetname, absolute_coordinate
# class management:
#   """
#   Switch sheet: Navigate between different sheets within the workbook.
#   Sort: Arrange the data within a range of cells in ascending or descending order.
#   Filter: Display only the rows of data that meet specified criteria.
#   Delete filter: Remove filter criteria from a range of cells.
#   Slicer: Create a slicer to filter data interactively.
#   Move rows: Change the position of rows within the spreadsheet.
#   Move columns: Change the position of columns within the spreadsheet.
#   Group: Group rows or columns together for organization.
#   Ungroup: Remove grouping from rows or columns.
#   Hide rows: Conceal rows from view.
#   Hide columns: Conceal columns from view.
#   Unhide rows: Reveal hidden rows.
#   Unhide columns: Reveal hidden columns.
#   Hide sheet: Make a sheet invisible within the workbook.
#   Unhide sheet: Make a previously hidden sheet visible again.
#   Set password: Protect the workbook or specific sheets with a password.
#   Transpose: Swap rows and columns in a range of cells.
#   Create named range: Define a named range of cells for easier reference.
#   Delete named range: Remove a named range from the workbook.
#   Data consolidation: Combine data from multiple ranges into a single range.
#   Freeze panes: Lock rows or columns to keep them visible while scrolling.
#   Unfreeze panes: Remove frozen panes from the spreadsheet.
#   Split panes: Divide the worksheet window into multiple resizable panes
#   """
#   def __init__(self):
#     return 


#   @staticmethod
#   def switch_sheet(workbook: excel.Workbook, sheet_name: str):
#     workbook.active = workbook[sheet_name]
#     return workbook
  
#   @staticmethod
#   def sort(workbook_path: str, source_sheet_name: str, key: str,output_workbook_name: str,ascending: bool = True):
#     wb = pd.read_excel(workbook_path,source_sheet_name)
#     if key in wb.columns:
#       df = wb.sort_values(by=key, ascending=ascending)
#     else:
#       return 
#     df.to_excel(output_workbook_name, source_sheet_name, index=False)
#   @staticmethod
#   def hide_null_value_rows(sheet, column, hidden):
#     column_index = column_index_from_string(column)
#     for row in sheet.iter_rows(min_row=2):
#       cell = row[column_index - 1]
#       if cell.value is None:
#         sheet.row_dimensions[row[0].row].hidden = hidden
#     return sheet
# #===================== in progress ============================
# #==============================================================  
#   #filter_criteria looks like this filter_criteria = {"key":"operator value" or "key":[start_range, end_range] or "key":"value" for equality or "key":"value*" for wildcard autocomplete}
#   @staticmethod
#   def filter(workbook_path: str, source_sheet_name: str,  key: str = "name", criteria = [], output_workbook_path: str = None):
#     # df = pd.read_feather(workbook_path,source_sheet_name)
#     # for operator, value in criteria:
#     #   if operator == ">":
#     #     df = df[df[key] > value]
#     #   elif operator == "<":
#     #     df = df[df[key] < value]
#     #   elif operator == "==":
#     #     df = df[df[key] == value]
#     #   elif operator == "<=":
#     #     df = df[df[key] <= value]
#     #   elif operator == ">=":
#     #     df = df[df[key] >= value]
#     #   elif operator == "!=":
#     #     df = df[df[key] != value]
#     #   elif operator == "contains":
#     #     df = df[df[key].str.contains(value)]
#     #   elif operator == "startswith":
#     #     df = df[df[key].str.startswith(value)]
#     #   elif operator == "endswith":
#     #     df = df[df[key].str.endswith(value)]
#     #   else:
#     #     raise ValueError(f"Unsupported operator: {operator}")
#     # df.to_excel(output_workbook_path)
#     wb = excel.load_workbook(workbook_path)
#     ws = wb[source_sheet_name]
#     names = [name[0].value.lower() for name in ws.columns]
#     # print(names)
#     index = names.index(key.lower())
#     key = get_column_letter(index + 1)
#     print(index, key)
#     ws.auto_filter.ref = f"{key}1:{key}{ws.max_row}"
#     ws.auto_filter.add_filter_column(index + 1,vals=[criteria], blank = False)
#     # ws = management.hide_null_value_rows(ws, key, True)
#     wb.save(output_workbook_path)
#     wb.close()
      
    
      
#   @staticmethod
#   def unfilter(workbook_path: str, source_sheet_name: str, output_workbook_path: str):
#     workbook = excel.load_workbook(workbook_path)
#     sheet = workbook[source_sheet_name]
#     sheet.auto_filter = None
#     workbook.save(output_workbook_path)
#     workbook.close()    
    
#   @staticmethod
#   def slicer(workbook_path: str, source_sheet_name: str, source_range: str, key: str, value: str):
#     return 
# #==============================================================
#   @staticmethod
#   def move_rows(workbook_path: str, source_sheet_name: str, row_range: str, new_pos: str):
#     workbook = excel.load_workbook(workbook_path)
#     worksheet = workbook[source_sheet_name]
#     old_position = range_boundaries(row_range)
#     new_pos = range_boundaries(new_pos)
#     end_pos = (new_pos[0] - old_position[0], new_pos[1] - old_position[1])
#     worksheet.move_range(row_range, end_pos[0], end_pos[1])
#     workbook.save(workbook_path)
#     workbook.close()
    
#   @staticmethod
#   def move_columns_by_name(workbook_path: str, source_sheet_name: str, column_name: str, new_pos: str):
#     workbook = excel.load_workbook(workbook_path)
#     worksheet = workbook[source_sheet_name]
#     names = [name[0].value.lower() for name in worksheet.columns]
#     index = names.index(column_name.lower())
#     key = get_column_letter(index)
#     old_pos = range_boundaries(f"{key}1")
#     new_pos = range_boundaries(new_pos)
#     end_pos = (new_pos[0] - old_pos[0], new_pos[1] - old_pos[1])
#     worksheet.move_range(f"{key}1:{key}{worksheet.max_row}", end_pos[0], end_pos[1])
#     workbook.save(workbook_path)
#     workbook.close()
  
#   @staticmethod
#   def move_columns_by_range(workbook_path: str, source_sheet_name: str, column_range: str, new_pos: str):
#     workbook = excel.load_workbook(workbook_path)
#     worksheet = workbook[source_sheet_name]
#     old_position = range_boundaries(column_range)
#     new_pos = range_boundaries(new_pos)
#     end_pos = (new_pos[0] - old_position[0], new_pos[1] - old_position[1])
#     column_range = column_range.split(":")
#     rng = f"{get_column_letter(old_position[0])}1:{get_column_letter(old_position[0])}{worksheet.max_row}"
#     worksheet.move_range(rng, end_pos[0], end_pos[1],True)
#     workbook.save(workbook_path)
#     workbook.close()
    
#   @staticmethod
#   def group(workbook_path: str, source_sheet_name: str, source_range: str, by: str, output_workbook_path: str,hidden: bool = True):
#     workbook = excel.load_workbook(workbook_path)
#     worksheet = workbook[source_sheet_name]
#     if source_range != None:
#       start_loc, end_loc = source_range.split(":")
#       start = range_boundaries(start_loc)
#       end = range_boundaries(end_loc)
#     elif source_range == None and by.lower() == "row":
#       start_loc = 2
#       end_loc = worksheet.max_row
#     elif source_range == None:
#       start_loc = 2
#       end_loc = worksheet.max_column
#     if by.lower() == "row":
#       worksheet.row_dimensions.group(start[1], end[3], hidden=hidden)
#     else:
#       worksheet.column_dimensions.group(start[0], end_loc[2], hidden=hidden)
#     workbook.save(output_workbook_path)
#     workbook.close()
    
#   @staticmethod
#   def ungroup(workbook_path: str, source_sheet_name: str, source_range: str, by: str, output_workbook_path: str):
#     wb = xls.Workbook()
#     wb.LoadFromFile(workbook_path)
#     sheet: xls.Worksheet = wb.Worksheets[source_sheet_name]
#     if source_range != "None":
#       start_loc, end_loc = source_range.split(":")
#       start = range_boundaries(start_loc)
#       end = range_boundaries(end_loc)
#     elif source_range == None and by.lower() == "row":
#       start_loc = 2
#       end_loc = 0
#     elif source_range == None:
#       start_loc = 1
#       end_loc = 0
#     if by.lower() == "row":
#       sheet.UngroupByRows(start[1], end[3])
#     else:
#       sheet.UngroupByColumns(start[0], end_loc[2])
#     sheet.UngroupByColumns(4, 6)
#     wb.SaveToFile(output_workbook_path, xls.ExcelVersion.Version2016)
#     wb.Dispose()
  
#   @staticmethod
#   def hide_rows(workbook_path: str, source_sheet_name: str, source_range: str):
#     workbook = excel.load_workbook(workbook_path)
#     worksheet = workbook[source_sheet_name]
#     for row in worksheet[source_range]:
#       for cell in row:
#         cell.hidden = True
#     workbook.save(workbook_path)
    
#   @staticmethod #needs adjustment
#   def hide_columns(workbook_path: str, source_sheet_name: str, source_range: str):
#     workbook = excel.load_workbook(workbook_path)
#     worksheet = workbook[source_sheet_name]
#     for row in worksheet[source_range]:
#       for cell in row:
#         cell.hidden = True
#     workbook.save(workbook_path)
  
#   @staticmethod 
#   def unhide_rows(workbook_path: str, source_sheet_name: str, source_range: str):
#     workbook = excel.load_workbook(workbook_path)
#     worksheet = workbook[source_sheet_name]
#     for row in worksheet[source_range]:
#       for cell in row:
#         cell.hidden = False
#     workbook.save(workbook_path)
    
#   @staticmethod #needs adjustment
#   def unhide_columns(workbook_path: str, source_sheet_name: str, source_range: str):
#     workbook = excel.load_workbook(workbook_path)
#     worksheet = workbook[source_sheet_name]
#     for row in worksheet[source_range]:
#       for cell in row:
#         cell.hidden = False
#     workbook.save(workbook_path)
    
#   @staticmethod
#   def hide_sheet(workbook_path: str, source_sheet_name: str):
#     workbook = excel.load_workbook(workbook_path)
#     worksheet = workbook[source_sheet_name]
#     worksheet.sheet_state = "hidden"
#     workbook.save(workbook_path)
    
    
#   @staticmethod
#   def unhide_sheet(workbook_path: str, source_sheet_name: str):
#     workbook = excel.load_workbook(workbook_path)
#     worksheet = workbook[source_sheet_name]
#     worksheet.sheet_state = "visible"
#     workbook.save(workbook_path)
  
#   @staticmethod
#   def set_password(workbook_path: str, source_sheet_name: str, password: str):
#     workbook = excel.load_workbook(workbook_path)
#     if source_sheet_name not in workbook.sheetnames:  
#       for sheet in workbook.worksheets:
#         sheet.protection.set_password(password)
#     else:
#       worksheet = workbook[source_sheet_name]
#       worksheet.protection.set_password(password)
#     workbook.save(workbook_path)
  
#   def transpose(workbook_path: str, source_sheet_name: str, source_range: str):
#     workbook = excel.load_workbook(workbook_path)
#     worksheet = workbook[source_sheet_name]
#     min_col, min_row, max_col, max_row = range_boundaries(source_range)
#     transposed_data = []
#     for row in range(min_row, max_row + 1):
#         transposed_row = []
#         for col in range(min_col, max_col + 1):
#             transposed_row.append(worksheet.cell(row=row, column=col).value)
#         transposed_data.append(transposed_row)

#     transposed_range = worksheet.iter_cols(min_row=min_row, min_col=min_col,
#                                        max_row=max_row, max_col=max_col)
#     for index, column in enumerate(transposed_range):
#         for idx, cell in enumerate(column):
#             cell.value = transposed_data[idx][index]
#     workbook.save(workbook_path)
    
#   @staticmethod
#   def create_named_range(workbook_path: str, source_sheet_name: str, range_name: str, source_range: str, output_workbook_path: str):
#     wb = excel.load_workbook(workbook_path)
#     ws = wb[source_sheet_name]
#     ref =  f"{quote_sheetname(ws.title)}!{absolute_coordinate(source_range)}"
#     defn = DefinedName(range_name, attr_text=ref)
#     wb.defined_names[range_name] = defn
#     wb.save(output_workbook_path)
#     wb.close()
    
    
#   @staticmethod
#   def delete_named_range(workbook_path: str, range_name: str, output_workbook_path: str):
#     wb = excel.load_workbook(workbook_path)
#     del wb.defined_names[range_name]
#     wb.save(output_workbook_path)
#     wb.close()

#   @staticmethod
#   def freeze_panes(workbook_path: str, sheet_name: str, top_left_cell: str, active_cell: str, output_workbook_path: str):
#     workbook = excel.load_workbook(workbook_path)
#     worksheet = workbook[sheet_name]
#     worksheet.freeze_panes = worksheet.cell(top_left_cell, active_cell)
#     workbook.save(output_workbook_path)
#     workbook.close()

#   @staticmethod
#   def unfreeze_panes(workbook_path: str, sheet_name: str,  output_workbook_path: str):
#     workbook = excel.load_workbook(workbook_path)
#     worksheet = workbook[sheet_name]
#     worksheet.freeze_panes = None
#     workbook.save(output_workbook_path)
#     workbook.close()
    
#   @staticmethod
#   def split_panes(workbook_path: str, sheet_name: str, vertical_split: str, horizontal_split: str, top_left_cell: str = 'A1', active_pane: str = "bottomRight"):
#     wb = xw.Workbook(workbook_path)
#     worksheet: xw.workbook.Worksheet = wb.get_worksheet_by_name(sheet_name)
#     worksheet.split_panes(vertical_split*10, horizontal_split*8.42)
#     wb.close()
    
#   @staticmethod
#   def data_consolidate(workbook_path: str, sheet_names: list[str], tgt_sheet: str = None):
#     worksheets = []
#     for sheet in sheet_names:
#       worksheets.append(pd.read_excel(workbook_path, sheet_name= sheet))
#     objs = pd.concat([worksheets], ignore_index = True)
#     if tgt_sheet != None:
#       with pd.ExcelWriter(workbook_path, engine='openpyxl', mode='a') as writer:
#           try:
#               objs.to_excel(writer, sheet_name = tgt_sheet, index=False, header=None)
#           except PermissionError:
#               print("Close the file in Excel and try again.")
      
    
    
    
# management.sort("./Modules/Action_Module/task_instructions.xlsx","Sheet1","No.","./Modules/Action_Module/task_instructions.xlsx",False)
# management.group("./Modules/Action_Module/task_instructions.xlsx","Sheet1","A1:B20","row","./Modules/Action_Module/task_instructions.xlsx",True)
# management.ungroup("./Modules/Action_Module/task_instructions.xlsx","Sheet1","A1:B20","row","./Modules/Action_Module/task_instructions.xlsx")
# management.filter("./Modules/Action_Module/task_instructions.xlsx","Sheet1","Categories", "Formatting","./Modules/Action_Module/task_instructions.xlsx")
# management.move_rows("./DemographicProfile"+".xlsx","Sheet1","A1:B20", "B20:B20")



from xlwings import constants as win32c
from constants import constants
import win32com.client as win32
from typing import List
from openpyxl.utils import get_column_letter
import os
import sys
script_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.append(script_dir)

class management_App():
  def __init__(self, app = 'excel', api_doc = None) -> None:
      super().__init__()
      self.appName = app
      self.__excel = None
      self.__currentWB = None

      if api_doc is not None:
          for key, value in api_doc.items():
              if value.get('display') is not None:
                  setattr(self, value['display'], self.__getattribute__(key))

  @property
  def activeAPP(self):
      if not self.__excel:
          try:
              self.__excel = win32.Dispatch('Excel.Application') if self.appName == 'excel' else win32.Dispatch('ket.Application')
              self.__excel.DisplayAlerts = False
              self.__excel.Visible = False
          except:
              raise Exception('{} is not running.'.format(self.appName))
      return self.__excel
  
  @property
  def activeWB(self):
      if self.__currentWB is not None:
          return self.__currentWB
      return self.activeAPP.ActiveWorkbook
  
  @activeWB.setter
  def activeWB(self, wb):
      self.__currentWB = wb
  
  @property
  def activeWS(self):
      return self.activeWB.ActiveSheet
  
  def toRangeUsingSheet(self, source: str):
      if '!' in source:
          sheet_name, source = source.split('!')
          sheet_name = sheet_name.strip("'") 
          if sheet_name not in [sheet.Name for sheet in self.activeWB.Worksheets]:
              raise ValueError(f'Sheet {sheet_name} does not exist.')
          sheet = self.activeWB.Sheets(sheet_name)
      else:
          raise Exception('The range must contain a sheet name.')
      
      if source.isdigit():
          return sheet.Rows(source)
      elif source.isalpha():
          return sheet.Columns(source)
      elif ':' in source or source.isalnum():
          return sheet.Range(source)
      
  def toRange(self, sheet,  source: str):
      if source.isdigit():
          return sheet.Rows(source)
      elif source.isalpha():
          return sheet.Columns(source)
      elif ':' in source or source.isalnum():
          return sheet.Range(source)
  
  def OpenWorkbook(self, path: str) -> None:
      self.__currentWB = self.activeAPP.Workbooks.Open(os.path.abspath(path))

  def SaveWorkbook(self, path: str) -> None:
      self.activeWB.SaveAs(os.path.abspath(path))
  
  def closeWorkBook(self) -> None:
      self.activeWB.Close()
     
  def Save(self) -> None:
    # self.activeWB.Save()
    pass
     
  def switch_sheet(self, sheet_name):
    if sheet_name not in self.activeWB.Sheets():
      return "did not switch sheet"
    name = self.activeWS.Name
    self.activeWS = self.activeWB.Sheets(sheet_name)
    return f"switched sheet from sheet named {name} to sheet named {sheet_name}"
  
  
  def sort(self, workbook_path, sheet_name, cell_range, order, orientation, output_workbook_path):
    # self.OpenWorkbook(workbook_path)
    sheet = self.activeWB.Sheets(sheet_name)
    source = self.toRange(sheet, cell_range)
    source.Sort(Key1=source, Order1=1 if order == 'asc' else 2, Orientation=1 if orientation == 'column' else 2)
    self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
    return f"sorted data in the range {cell_range} using {order} order and the {orientation} orientation in sheet {sheet_name}"
  
  def filter(self, workbook_path, sheet_name, cell_range, feild_index, criteria, output_workbook_path):
    # self.OpenWorkbook(workbook_path)
    sheet = self.activeWB.Sheets(sheet_name)
    source = self.toRange(sheet, cell_range)
    try:
        criteriaRange = self.toRange(criteria)
    except:
        criteriaRange = None
    if criteriaRange:
        criteria = [criteriaRange.Cells(i).Text for i in range(1, criteriaRange.Cells.Count + 1)]
    source.AutoFilter(Field=feild_index, Criteria1=criteria)
    self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
    return f"applied filter on column {feild_index} using condition {criteria} in sheet {sheet_name}"
  
  def delete_filter(self, workbook_path, sheet_name, output_workbook_path):
    # self.OpenWorkbook(workbook_path)
    sheet = self.activeWB.Sheets(sheet_name)
    if sheet.AutoFilterMode: self.activeWS.AutoFilterMode = False
    self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
    return f"deleted filters on sheet {sheet_name}"
  
  def slicer(self, workbook_path: str, source_sheet_name: str, source_range: str, key: str, value: str, output_workbook_path: str):
    # self.OpenWorkbook(workbook_path)
    sheet = self.activeWB.Sheets(source_sheet_name)
    
    try:
        range_to_slice = self.toRange(sheet, source_range)
        slicer_cache = self.activeWB.SlicerCaches.Add2(range_to_slice, key)
        slicer = slicer_cache.Slicers.Add(sheet, key, value)
    except Exception as e:
        print(f"Error: {e}")

    self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
    return f"Created slicer on range {source_range} with key {key} and value {value} in sheet {source_sheet_name}"

  def move_rows(self, workbook_path: str, source_sheet_name: str, row_range: tuple, new_pos: int, output_workbook_path):
    # self.OpenWorkbook(workbook_path)
    sheet = self.activeWB.Sheets(source_sheet_name)
    source_start = row_range[0]
    source_end = row_range[1]
    num_rows = abs(source_end - source_start + 1)
    if new_pos >= source_start and new_pos <= source_end:
        raise ValueError("Destination range overlaps with source range.")
    for i in range(num_rows):
        sheet.Rows(new_pos + num_rows).Insert(Shift=win32.constants.xlShiftDown)
    # Copy the source rows to the destination
    source_range = sheet.Rows(f"{source_start}:{source_end}")
    dest_range = sheet.Rows(new_pos + num_rows)
    source_range.Copy(dest_range)

    # Delete the original source rows
    source_range.Delete(Shift=win32.constants.xlShiftUp)
    self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
    return f"moved rows {row_range[0]} - {row_range[1]} to a new position at {new_pos} in sheet '{source_sheet_name}'"
 
  def move_columns(self, workbook_path: str, source_sheet_name: str, column_range: tuple, new_pos: int, output_workbook_path: str):
    # self.OpenWorkbook(workbook_path)
    sheet = self.activeWB.Sheets(source_sheet_name)
    start_col, end_col = column_range
    num_cols = abs(end_col - start_col + 1)

    if new_pos >= start_col and new_pos <= end_col:
        raise ValueError("Destination range overlaps with source range.")

    # Adjust new position to avoid overlap
    if new_pos > end_col:
        new_pos += 1

    # Insert blank columns at the new position
    for _ in range(num_cols):
        sheet.Columns(new_pos).Insert(Shift=win32.constants.xlShiftToRight)

    # Copy the source columns to the new position
    start_col_letter = get_column_letter(start_col)
    end_col_letter = get_column_letter(end_col)
    new_pos_letter = get_column_letter(new_pos)
    source_range = self.toRange(sheet, f"{start_col_letter}:{end_col_letter}")
    dest_range = self.toRange(sheet, new_pos_letter)
    source_range.Copy(dest_range)

    # Delete the original source columns
    source_range.Delete(Shift=win32.constants.xlShiftToLeft)

    self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()

    return f"Moved columns {start_col} - {end_col} to a new position at {new_pos} in sheet '{source_sheet_name}'"

  def group(self, workbook_path: str, source_sheet_name: str, source_range: str, output_workbook_path: str, group_by_rows: bool = True,  hidden: bool = True):
    # self.OpenWorkbook(workbook_path)
    try:
        sheet = self.activeWB.Sheets(source_sheet_name)
        group_range = self.toRange(sheet, source_range)
        if group_by_rows:
            group_range.Rows.Group()
            group_range.Rows.Hidden = hidden
        else:
            group_range.Columns.Group()
            group_range.Columns.Hidden = hidden
    except Exception as e:
        print(f"Error: {e}")
    self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
    return f"grouped range {source_range} in sheet '{source_sheet_name}'"
  
  def ungroup(self, workbook_path: str, source_sheet_name: str, source_range: str, output_workbook_path: str, group_by_rows: bool = True):
    # self.OpenWorkbook(workbook_path)
    try:
        sheet = self.activeWB.Sheets(source_sheet_name)
        group_range = self.toRange(sheet, source_range)
        if group_by_rows:
            group_range.Rows.Ungroup()
        else:
            group_range.Columns.Ungroup()
    except Exception as e:
        print(f"Error: {e}")
    self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
    return f"ungrouped range {source_range} in sheet '{source_sheet_name}'"
  
  def hide_unhide_rows(self, workbook_path: str, source_sheet_name: str, start_row, end_row, output_workbook_path: str, hidden = True):
    # self.OpenWorkbook(workbook_path)
    sheet = self.activeWB.Sheets(source_sheet_name)
    group_range = sheet.Rows(f"{start_row}:{end_row}")
    try:
      group_range.Hidden = hidden
        # Save the workbook
    except Exception as e:
        print(f"Error: {e}")
    self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
    return f"{'hide' if hidden else 'unhide'} rows({start_row} - {end_row}) in sheet '{source_sheet_name}'"
  
  def hide_unhide_columns(self, workbook_path: str, source_sheet_name: str, start_col: int, end_col: int, output_workbook_path: str, hidden: bool = True):
    # self.OpenWorkbook(workbook_path)
    sheet = self.activeWB.Sheets(source_sheet_name)
    start_col_letter = get_column_letter(start_col)
    end_col_letter = get_column_letter(end_col)
    group_range = sheet.Range(f"{start_col_letter}:{end_col_letter}")
    try:
        group_range.EntireColumn.Hidden = hidden
    except Exception as e:
        print(f"Error: {e}")
    self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
    return f"{'hide' if hidden else 'unhide'} columns({start_col} - {end_col}) in sheet '{source_sheet_name}'"
 
  def hide_unhide_sheet(self, workbook_path: str, source_sheet_name: str, hidden: bool, output_workbook_path: str):
    # self.OpenWorkbook(workbook_path)
    sheet = self.activeWB.Sheets(source_sheet_name)
    try:
        sheet.Visible = 0 if hidden else -1
    except Exception as e:
        print(f"Error: {e}")
    self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
    return f"{'hide' if hidden else 'unhide'} sheet '{source_sheet_name}'"
  def set_password(self, workbook_path: str, source_sheet_name: str, password: str, output_workbook_path, wb = False):
    # self.OpenWorkbook(workbook_path)
    sheet = self.activeWB.Sheets(source_sheet_name)
    try:
      if wb:
        self.activeWB.Password = password
      else:
        sheet.Protect(Password = password)
    except Exception as e:
      print(f"Error: {e}")
    self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
    return f"{'workbook is' if wb else f'sheet {source_sheet_name} is'}  protected using password {['*'] * len(password)}"
  
  def transpose(self, workbook_path: str, source_sheet_name: str, source_range: str, output_workbook_path: str):
    # self.OpenWorkbook(workbook_path)
    sheet = self.activeWB.Sheets(source_sheet_name)
    source = self.toRange(sheet, source_range)  # Note: `toRange` method should not require `sheet` as an argument.
    dataT = self.activeAPP.WorksheetFunction.Transpose(source)
    source.Clear()
    
    # Determining the starting cell for the transposed data
    start_row = source.Row
    start_col = source.Column
    # print(start_row, start_col)
    # print(dataT)
    if isinstance(dataT, list) and isinstance(dataT[0], list):
        for rowOffset, row in enumerate(dataT):
            for colOffset, value in enumerate(row):
                sheet.Cells(start_row + rowOffset, start_col + colOffset).Value = value
    else:
        for rowOffset, row in enumerate(list(dataT)):
          for column_offset, column in enumerate(list(row)):
            sheet.Cells(start_row + rowOffset, start_col + column_offset).Value = column
    
    self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
    return f"transposed the range {source_range} in sheet '{source_sheet_name}'"

  def create_named_range(self, workbook_path: str, source_sheet_name: str,  range_name: str, source_range: str, output_workbook_path: str):
    # self.OpenWorkbook(workbook_path)
    sheet = self.activeWB.Sheets(source_sheet_name)
    source = self.toRange(sheet, source_range)
    try:
      source.Name = range_name
    except Exception as e:
      print(f"Error: {e}")
    self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
    return f"transposed the range {source_range} in sheet '{source_sheet_name}'"
  
  def freeze_panes(self, workbook_path: str, sheet_name: str, range_obj: str, output_workbook_path: str):
    # self.OpenWorkbook(workbook_path)
    sheet = self.activeWB.Sheets(sheet_name)
    source = self.toRange(sheet, range_obj)
    source.Select()
    source.Parent.Application.ActiveWindow.FreezePanes = True
    self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
    return f"froze panes range {range_obj} in sheet '{sheet_name}'"
  
  def unfreeze_panes(self, workbook_path: str, sheet_name: str,  output_workbook_path: str):
    # self.OpenWorkbook(workbook_path)
    sheet = self.activeWB.Sheets(sheet_name)
    if sheet_name:
        sheet = self.activeWB.Worksheets(sheet_name)
    else:
        sheet = self.activeWS
    sheet.Application.ActiveWindow.FreezePanes = False
    self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
    return f"unfroze panes in sheet '{sheet_name}'"
  
  def split_panes(self, workbook_path: str, sheet_name: str, vertical_split: str, horizontal_split: str,  output_workbook_path: str):
    # self.OpenWorkbook(workbook_path)
    sheet = self.activeWB.Sheets(sheet_name)
    try:
      self.activeAPP.ActiveWindow.SplitRow = vertical_split
      self.activeAPP.ActiveWindow.SplitColumn = horizontal_split
    except Exception as e:
      print(f"Error: {e}")
    self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
    return f"unfroze panes in sheet '{sheet_name}'"
  
  def data_consolidation(self, workbook_path: str, source_ranges: list, destination_sheet: str, destination_range: str, output_workbook_path: str, function: int = 1, top_row: bool = False, left_column: bool = False, create_links: bool = False):
    # self.OpenWorkbook(workbook_path)
    sheet = self.activeWB.Sheets(destination_sheet)
    
    try:
        # Get the destination range
        dest_range = sheet.Range(destination_range)
        
        # Perform data consolidation
        dest_range.Consolidate(Sources=source_ranges, Function=function, TopRow=top_row, LeftColumn=left_column, CreateLinks=create_links)
    except Exception as e:
        print(f"Error: {e}")
    
    self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
    
    return f"Data consolidated for ranges {source_ranges} in sheet '{destination_sheet}' into range '{destination_range}'"

    
    
    
    
# app = management_App() <== works
# app.sort("./DemographicProfile.xlsx", "Sheet1", "A1:A4", "desc", "column", "./DemographicProfile.xlsx") <== works
# app.filter("./DemographicProfile.xlsx", "Sheet1", "A1:A4", "desc", "column", "./DemographicProfile.xlsx") <== test
# app.delete_filter() <== test after testing filter
# app.slicer("./DemographicProfile.xlsx", "Sheet1", "A1:A4", "E", "Elemen", "./DemographicProfile.xlsx") <== fix and retest
# app.move_rows("./DemographicProfile.xlsx", "Sheet1", (1, 3), 20, "./DemographicProfile.xlsx") <== works
# app.move_columns("./DemographicProfile.xlsx", "Sheet1", (3, 4), 5, "./DemographicProfile.xlsx") <== retest later
# app.group("./DemographicProfile.xlsx", "Sheet1", "3:10", "./DemographicProfile.xlsx") <== works
# app.ungroup("./DemographicProfile.xlsx", "Sheet1", "3:10", "./DemographicProfile.xlsx") <== works
# app.hide_unhide_rows("./DemographicProfile.xlsx", "Sheet1", 2, 5, "./DemographicProfile.xlsx",False) <== works
# app.hide_unhide_columns("./DemographicProfile.xlsx", "Sheet1", 3, 4, "./DemographicProfile.xlsx", False) <== works
# app.hide_unhide_sheet("./DemographicProfile.xlsx", "Sheet1", False, "./DemographicProfile.xlsx") <== works
# app.transpose("./DemographicProfile.xlsx", "Sheet1", "A1:D41", "./DemographicProfile.xlsx") <== works
# app.freeze_panes("./DemographicProfile.xlsx", "Sheet1", "A1:D1", "./DemographicProfile.xlsx") <== works
# app.unfreeze_panes("./DemographicProfile.xlsx", "Sheet1", "./DemographicProfile.xlsx") <== works
# app.data_consolidation("./DemographicProfile.xlsx", ["Sheet1!A1:B10", "Sheet2!C1:D10"], "Sheet1", "A1:B10", "./DemographicProfile.xlsx") <== fix later
