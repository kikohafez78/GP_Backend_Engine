# from openpyxl import load_workbook
# from openpyxl.drawing.image import Image
# from openpyxl.utils import get_column_letter
# from openpyxl.utils.cell import coordinate_from_string, range_boundaries
# from openpyxl.utils import cell as cs
# import xlwings as xw

# class entryandmanipulation(object):
#     def __init__(self):
#         super().__init__()
    
#     @staticmethod
#     def update_cell_value(workbook_path, sheet_name, cell_range, new_value, output_workbook_path):
#         wb = load_workbook(workbook_path)
#         sheet = wb[sheet_name]
#         position = range_boundaries(cell_range)
#         for row in range(position[1], position[3] + 1):
#             for col in range(position[0], position[2] + 1):
#                 sheet.cell(row,col).value = new_value
#         wb.save(output_workbook_path)
#         wb.close()

#     @staticmethod
#     def delete_cells(workbook_path, sheet_name, cell_range,output_workbook_path):
#         wb = load_workbook(workbook_path)
#         sheet = wb[sheet_name]
#         start_col, start_row, end_col, end_row = range_boundaries(cell_range)
#         if start_col == end_col:
#             sheet.delete_rows(start_row, end_row - start_row + 1)
#         elif start_row == end_row:
#             sheet.delete_cols(start_col, end_col - start_col + 1)
#         else:
#             sheet.delete_rows(start_row, end_row - start_row + 1)
#             sheet.delete_cols(start_col, end_col - start_col + 1)
#         wb.save(output_workbook_path)
#         wb.close()

#     @staticmethod
#     def merge_cells(workbook_path, sheet_name, cell_range, output_workbook_path):
#         wb = load_workbook(workbook_path)
#         sheet = wb[sheet_name]
#         sheet.merge_cells(cell_range)
#         wb.save(output_workbook_path)
#         wb.close()

#     @staticmethod
#     def unmerge_cells(workbook_path, sheet_name, cell_range, output_workbook_path):
#         wb = load_workbook(workbook_path)
#         sheet = wb[sheet_name]
#         sheet.unmerge_cells(cell_range)
#         wb.save(output_workbook_path)
#         wb.close()
    
#     @staticmethod
#     def split_text_to_columns(workbook_path, sheet_name, cell_range, output_workbook_path: str = "", delimiter:str = "."):
#         wb = load_workbook(workbook_path)
#         sheet = wb[sheet_name]
#         start_col, start_row, end_col, end_row = range_boundaries(cell_range)
#         max_len = 0
#         for i in range(start_row, end_row + 1):
#             for j in range(start_col, end_col + 1):
#                 cell = sheet.cell(row=i, column=j) 
#                 max_len = max(max_len, len(cell.value.split(delimiter)))
#         ranges = f"{get_column_letter(end_col + 1)}1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
#         sheet.move_range(ranges,0,max_len - 1)
#         for i in range(start_row, end_row + 1):
#             for j in range(start_col, end_col + 1):
#                 cell = sheet.cell(row=i, column=j)
#                 string = cell.value.split(delimiter) 
#                 for k, value in enumerate(string):
#                         sheet.cell(row=i, column=j + k).value = value
#         wb.save(output_workbook_path)
#         wb.close()

#     @staticmethod
#     def insert_row(workbook_path, sheet_name, row_number, output_workbook_path:str = ""):
#         wb = load_workbook(workbook_path)
#         sheet = wb[sheet_name]
#         ranges = f"A{row_number + 1}:{get_column_letter(sheet.max_column)}{sheet.max_row}"
#         sheet.move_range(ranges,0,1)
#         sheet.insert_rows(row_number)
#         wb.save(output_workbook_path)
#         wb.close()

#     @staticmethod
#     def insert_column(workbook_path, sheet_name, column_number, column_name: str = "new column", output_workbook_path:str = ""):
#         wb = load_workbook(workbook_path)
#         sheet = wb[sheet_name]
#         ranges = f"{get_column_letter(column_number)}1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
#         sheet.move_range(ranges,0,1)
#         sheet.insert_cols(column_number)
#         sheet[get_column_letter(column_number) + "1"].value = column_name
#         wb.save(output_workbook_path)
#         wb.close()

#     @staticmethod
#     def autofill(workbook_path, sheet_name, cell_range,formula_cell: str, output_workbook_path):
#         app = xw.App(visible = False)
#         workbook  = app.books.open(workbook_path)
#         sheet = workbook.sheets[0]
#         rng = sheet.range(cell_range)
#         sheet.range(cell_range).api.Autofill(sheet.range(formula_cell).api, 0)
#         workbook.save(output_workbook_path)
#         workbook.close()
#         app.quit()


#     @staticmethod
#     def copy_paste_range(source_workbook_path: str, target_workbook_path: str,source_sheet_name: str, target_sheet_name: str, source_range: str, target_cell: str):
#         src_wb = load_workbook(source_workbook_path)
#         tgt_wb = load_workbook(target_workbook_path)
#         src_sheet = src_wb[source_sheet_name]
#         tgt_sheet = tgt_wb[target_sheet_name]
#         start_column, start_row, end_column, end_row = range_boundaries(source_range)
#         position = range_boundaries(target_cell)
#         print(start_column,start_row,end_column,end_row)
#         print(position)
#         for i in range(start_row, end_row + 1):
#             for j in range(start_column, end_column + 1):
#                 cell = src_sheet.cell(row=i, column=j)
#                 tgt_sheet.cell(row = position[1] + i - start_row, column = position[0] + j - start_column).value = cell.value
#         tgt_wb.save(target_workbook_path)
#         tgt_wb.close()
#         src_wb.close()

#     @staticmethod
#     def copy_paste_format(source_workbook_path: str, target_workbook_path: str,source_sheet_name: str, target_sheet_name: str, source_range: str, target_cell: str):
#         src_wb = load_workbook(source_workbook_path)
#         tgt_wb = load_workbook(target_workbook_path)
#         src_sheet = src_wb[source_sheet_name]
#         tgt_sheet = tgt_wb[target_sheet_name]
#         start_column, start_row, end_column, end_row = range_boundaries(source_range)
#         position = range_boundaries(target_cell)
#         print(start_column,start_row,end_column,end_row)
#         print(position)
#         for i in range(start_row, end_row + 1):
#             for j in range(start_column, end_column + 1):
#                 cell = src_sheet.cell(row=i, column=j)
#                 tgt_sheet.cell(row = position[1] + i - start_row, column = position[0] + j - start_column).value = cell.value
#                 if cell.has_style:
#                     tgt_sheet.cell(row = position[1] + i - start_row, column = position[0] + j - start_column).style = cell.style    
#         tgt_wb.save(target_workbook_path)
#         tgt_wb.close()
#         src_wb.close()

#     @staticmethod
#     def copy_sheet(source_wb_file, dest_wb_file, sheet_name, new_sheet_name="new Sheet"):
#         src_wb = load_workbook(source_wb_file)
#         tgt_wb = load_workbook(dest_wb_file)
#         src_sheet = src_wb[sheet_name]
#         tgt_wb.create_sheet(new_sheet_name)
#         entryandmanipulation.copy_paste_format(source_wb_file, dest_wb_file,sheet_name,new_sheet_name,f"A1:{get_column_letter(src_sheet.max_column)}{src_sheet.max_row}","A1")
#         src_wb.close()
#         tgt_wb.close()


#     @staticmethod
#     def cut_paste_range(src_workbook_path, tgt_workbook_path, src_sheet_name,tgt_sheet_name, source_range, target_range):
#         entryandmanipulation.copy_paste_range(src_workbook_path,tgt_workbook_path, src_sheet_name, tgt_sheet_name,  source_range, target_range)
#         entryandmanipulation.delete_cells(src_workbook_path, src_sheet_name, source_range)
        
        
#     @staticmethod
#     def find_and_replace(workbook_path, sheet_name, find_text, replace_text):
#         wb = load_workbook(workbook_path)
#         sheet = wb[sheet_name]
#         if type(find_text) == str:
#             for row in sheet:
#                 for cell in row:
#                     if find_text in cell.value:
#                         cell.value.replace(find_text, replace_text)
#         else:
#             for row in sheet:
#                 for cell in row:
#                     if find_text == cell.value:
#                         cell.value = replace_text
#         wb.save(workbook_path)
#         wb.close()

#     @staticmethod
#     def set_hyperlink(workbook_path, sheet_name, cell_range, url):
#         wb = load_workbook(workbook_path)
#         sheet = wb[sheet_name]
#         for row in sheet[cell_range]:
#             for cell in row:
#                 cell.hyperlink = url
#         wb.save(workbook_path)

#     @staticmethod
#     def delete_hyperlink(workbook_path, sheet_name, cell_range):
#         wb = load_workbook(workbook_path)
#         sheet = wb[sheet_name]
#         for row in sheet[cell_range]:
#             for cell in row:
#                 if cell.hyperlink:
#                     cell.hyperlink = None
#         wb.save(workbook_path)

#     @staticmethod
#     def remove_duplicates(workbook_path, sheet_name, column_number):
#         wb = load_workbook(workbook_path)
#         sheet = wb[sheet_name]
#         values = set()
#         position = get_column_letter(column_number)
#         for row in sheet[f"{position}2:{position}{sheet.max_row}"]:
#             for cell in row:
#                 if cell.value in values:
#                     sheet.delete_rows(cell.row)
#                 else:
#                     values.add(cell.value)
#         wb.save(workbook_path)

#     @staticmethod
#     def rename_sheet(workbook_path, old_sheet_name, new_sheet_name):
#         wb = load_workbook(workbook_path)
#         wb[old_sheet_name].title = new_sheet_name
#         wb.save(workbook_path)

#     @staticmethod
#     def insert_checkbox(workbook_path, sheet_name, cell_range):
#         wb = load_workbook(workbook_path)
#         sheet = wb[sheet_name]
#         sheet.cell(*cell_range).value = '☑'
#         wb.save(workbook_path)
#         wb.close()

#     @staticmethod #needs work
#     def insert_textbox(workbook_path, sheet_name, cell_range, text):
#         wb = load_workbook(workbook_path)
#         sheet = wb[sheet_name]
#         img = Image('path/to/your/textbox_image.png')
#         img.anchor = cell_range
#         sheet.add_image(img)
#         wb.save(workbook_path)
#         wb.close()

#     @staticmethod
#     def create_sheet(workbook_path, sheet_name):
#         wb = load_workbook(workbook_path)
#         wb.create_sheet(title = sheet_name)
#         wb.save(workbook_path)
#         wb.close()

#     @staticmethod
#     def delete_sheet(workbook_path, sheet_name):
#         wb = load_workbook(workbook_path)
#         del wb[sheet_name]
#         wb.save(workbook_path)
#         wb.close()

#     @staticmethod
#     def clear_range(workbook_path, sheet_name, cell_range):
#         wb = load_workbook(workbook_path)
#         sheet = wb[sheet_name]
#         for row in sheet[cell_range]:
#             for cell in row:
#                 cell.value = None  # Clear cell value
#                 # if cell.has_style:
#                 #     cell.style = None# Remove any formatting
#         wb.save(workbook_path)
#         wb.close()
    





#testing
# entryandmanipulation.update_cell_value("./Modules/Action_Module/testing.xlsx","Sheet1","A1:B20","hello.world.kiko","./Modules/Action_Module/testing.xlsx")
# entryandmanipulation.delete_cells("./Modules/Action_Module/testing.xlsx","Sheet1","A1:B20","./Modules/Action_Module/testing.xlsx")
# entryandmanipulation.merge_cells("./Modules/Action_Module/testing.xlsx","Sheet1","A1:B10","./Modules/Action_Module/testing.xlsx")
# entryandmanipulation.split_text_to_columns("./Modules/Action_Module/testing.xlsx","Sheet1","A1:A20","./Modules/Action_Module/testing.xlsx")
# entryandmanipulation.autofill("./Modules/Action_Module/testing.xlsx", "Sheet1","A1:A20","./Modules/Action_Module/testing.xlsx")
# entryandmanipulation.insert_column("./Modules/Action_Module/seed_tasks.xlsx","Sheet1",2,"cash","./Modules/Action_Module/seed_tasks.xlsx")
# entryandmanipulation.insert_row("./Modules/Action_Module/seed_tasks.xlsx","Sheet1",2,"./Modules/Action_Module/seed_tasks.xlsx")
# entryandmanipulation.copy_paste_format("./Modules/Action_Module/task_instructions.xlsx","./Modules/Action_Module/task_instructions.xlsx","Sheet1","Sheet1","A1:A20","B1")
# entryandmanipulation.copy_paste_range("./Modules/Action_Module/task_instructions.xlsx","./Modules/Action_Module/task_instructions.xlsx","Sheet1","Sheet1","A1:A20","B1")
# entryandmanipulation.remove_duplicates("./Modules/Action_Module/task_instructions.xlsx","Sheet1", 6)
# entryandmanipulation.delete_hyperlink("./Modules/Action_Module/task_instructions.xlsx","Sheet1","A1:A20")
# entryandmanipulation.create_sheet("./Modules/Action_Module/task_instructions.xlsx","Sheet2")
# entryandmanipulation.delete_sheet("./Modules/Action_Module/task_instructions.xlsx","Sheet2")
# entryandmanipulation.find_and_replace("./Modules/Action_Module/task_instructions.xlsx","Sheet1",80,180)
# entryandmanipulation.clear_range("./Modules/Action_Module/task_instructions.xlsx","Sheet1","A1:A20")








import win32com.client as win32
import pythoncom
from typing import List
from openpyxl.utils import get_column_letter
import os
import sys
import time
script_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.append(script_dir)

class entry_manipulation_App():
    def __init__(self, app = 'excel', api_doc = None) -> None:
        super().__init__()
        self.appName = app
        self.__excel = None
        self.__currentWB = None

        if api_doc is not None:
            for key, value in api_doc.items():
                if value.get('display') is not None:
                    setattr(self, value['display'], self.__getattribute__(key))

    @property
    def activeAPP(self):
        if not self.__excel:
            try:
                pythoncom.CoInitialize()
                self.__excel = win32.Dispatch('Excel.Application') if self.appName == 'excel' else win32.Dispatch('ket.Application')
                self.__excel.DisplayAlerts = False
                self.__excel.Visible = True
            except:
                raise Exception('{} is not running.'.format(self.appName))
        return self.__excel
    
    @property
    def activeWB(self):
        if self.__currentWB is not None:
            return self.__currentWB
        return self.activeAPP.ActiveWorkbook
    
    @activeWB.setter
    def activeWB(self, wb):
        self.__currentWB = wb
    
    @property
    def activeWS(self):
        return self.activeWB.ActiveSheet
    
    def toRangeUsingSheet(self, source: str):
        if '!' in source:
            sheet_name, source = source.split('!')
            sheet_name = sheet_name.strip("'") 
            if sheet_name not in [sheet.Name for sheet in self.activeWB.Worksheets]:
                raise ValueError(f'Sheet {sheet_name} does not exist.')
            sheet = self.activeWB.Sheets(sheet_name)
        else:
            raise Exception('The range must contain a sheet name.')
        
        if source.isdigit():
            return sheet.Rows(source)
        elif source.isalpha():
            return sheet.Columns(source)
        elif ':' in source or source.isalnum():
            return sheet.Range(source)
        
    def toRange(self, sheet,  source: str):
        if source.isdigit():
            return sheet.Rows(source)
        elif source.isalpha():
            return sheet.Columns(source)
        elif ':' in source or source.isalnum():
            return sheet.Range(source)
    
    def OpenWorkbook(self, path: str) -> None:
        self.__currentWB = self.activeAPP.Workbooks.Open(os.path.abspath(path))

    def SaveWorkbook(self, path: str) -> None:
        self.activeWB.SaveAs(os.path.abspath(path))
    
    def closeWorkBook(self) -> None:
        self.activeWB.Close()
    
    def Save(self) -> None:
        '''
        Saves the current workbook in its existing location.

        Returns:
        None
        '''
        # self.activeWB.Save()
        pass
        
    def update_cell_value(self, workbook_path: str, destination_sheet: str, cell_range: str, value, output_workbook_path: str):
        # self.OpenWorkbook(workbook_path)
        if destination_sheet not in [sheet.Name for sheet in self.activeWB.Worksheets]:
            raise ValueError(f'Sheet {destination_sheet} does not exist.')
        sheet = self.activeWB.Sheets(destination_sheet)
        destination_range = self.toRange(sheet, cell_range)
        if isinstance(value, (list, tuple)) and range.Count == 1:
            if isinstance(value[0], (list, tuple)):
                for rowOffet, elem in enumerate(value):
                    for columnOffset, elem2 in enumerate(elem):
                        destination_range.GetOffset(rowOffet, columnOffset).Value = elem2
            else:
                for columnOffset, elem in enumerate(value):
                    destination_range.GetOffset(0, columnOffset).Value = elem
        else:
            
            destination_range.Value = value
        self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
        return f"update cell range {cell_range} in sheet {destination_sheet} with value {value}"
    
    def delete_cells(self, workbook_path, sheet_name, cell_range, output_workbook_path):
        # self.OpenWorkbook(workbook_path)
        sheet = self.activeWB.Sheets(sheet_name)
        cell_range_to_del = self.toRange(sheet, cell_range)
        cell_range_to_del.Delete()
        self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
        return f"deleted cells in the range {cell_range} for sheet {sheet_name}"
    
    
    def merge_cells(self, workbook_path: str, sheet_name, cell_range, output_workbook_path: str):
        # self.OpenWorkbook(workbook_path)
        sheet = self.activeWB.Sheets(sheet_name)
        cells_to_merge = self.toRange(sheet, cell_range)
        cells_to_merge.Merge()
        self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook() 
        return f"merged cells in the range {cell_range}"
    
    def unmerge_cells(self, workbook_path: str, sheet_name, cell_range, output_workbook_path: str):
        # self.OpenWorkbook(workbook_path)
        sheet = self.activeWB.Sheets(sheet_name)
        cells_to_merge = self.toRange(sheet, cell_range)
        cells_to_merge.UnMerge()
        self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook() 
        return f"Unmerged cells in the range {cell_range}"
    
    def InsertColumn(self, column: int) -> None:
        self.activeWS.Columns(column).Insert()
    
    
    def split_text_to_columns(self, workbook_path, sheet_name, cell_range, output_workbook_path: str = "", delimiter: str = "."):
        # self.OpenWorkbook(workbook_path)
        sheet = self.activeWB.Sheets(sheet_name)
        source = self.toRange(sheet, cell_range)
        
        if source.Columns.Count != 1:
            print('Source must be one column.')
            return

        orgData = source.Value
        if isinstance(orgData, tuple):
            orgData = [cell[0] for cell in orgData]

        newData = [x.split(delimiter) for x in orgData]
        maxLen = max(len(x) for x in newData)
        newData = [x + [None] * (maxLen - len(x)) for x in newData]

        for i in range(maxLen - 1):
            self.InsertColumn(source.Column + i + 1)

        for rowOffset, row in enumerate(newData):
            for colOffset, value in enumerate(row):
                sheet.Cells(source.Row + rowOffset, source.Column + colOffset).Value = value

        self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
        return f"Splitted text cells in the range {cell_range} on the delimiter '{delimiter}'."


    def insert_columns(self, workbook_path, sheet_name, before_column, after_column, count, output_workbook_path):
        # self.OpenWorkbook(workbook_path)
        sheet = self.activeWB.Sheets(sheet_name)
        if before_column:
            index = before_column
        elif after_column:
            index = after_column + 1
        else:
            raise ValueError("Either before_column or after_column must be specified.")
        # Insert the required number of blank columns
        for _ in range(count):
            sheet.Columns(index).Insert(Shift=win32.constants.xlShiftToRight)
        self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook() 
        return f"inserted new column at position {after_column} between {before_column} and {after_column} in the sheet {sheet_name}"
    
    def insert_rows(self, workbook_path, sheet_name, above_row, below_row, count, output_workbook_path):
        # self.OpenWorkbook(workbook_path)
        sheet = self.activeWB.Sheets(sheet_name)
        
        if above_row:
            index = above_row
        elif below_row:
            index = below_row + 1
        else:
            raise ValueError("Either 'above_row' or 'below_row' must be specified.")

        for i in range(count):
            sheet.Rows(index).Insert(Shift=win32.constants.xlShiftDown)

        self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
        
        return f"Inserted {count} new row(s) at position {index} in the sheet {sheet_name}"

    def autofill(self, workbook_path, sheet_name, cell_range, dest_sheet, dest_range, output_workbook_path):
        # self.OpenWorkbook(workbook_path)
        sheet = self.activeWB.Sheets(sheet_name)
        dest_sheet = self.activeWB.Sheets(dest_sheet)
        source = self.toRange(sheet, cell_range)
        destination = self.toRange(dest_sheet, dest_range)
        source.AutoFill(destination)
        self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook() 
        return f"autofilled {cell_range} in sheet name {sheet_name}"
    
    def copy_paste(self, source_workbook_path: str, source_sheet_name: str, target_sheet_name: str, source_range: str, target_range: str, output_workbook_path: str):
        self.OpenWorkbook(source_workbook_path)
        sheet_1 = self.activeWB.Sheets(source_sheet_name)
        sheet_2 = self.activeWB.Sheets(target_sheet_name)
        src_range = self.toRange(sheet_1, source_range).SpecialCells(12)
        dest_range = self.toRange(sheet_2, target_range)
        src_range.Copy()
        dest_range.PasteSpecial(-4163)
        self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook() 
        return f"copied {source_range} in sheet name {source_sheet_name} to {target_range} in sheet name {target_sheet_name}"
    
    def copy_paste_format(self, source_workbook_path: str, source_sheet_name: str, target_sheet_name: str, source_range: str, target_range: str, output_workbook_path: str):
        self.OpenWorkbook(source_workbook_path)
        sheet_1 = self.activeWB.Sheets(source_sheet_name)
        sheet_2 = self.activeWB.Sheets(target_sheet_name)
        src_range = self.toRange(sheet_1, source_range)
        dest_range = self.toRange(sheet_2, target_range)
        src_range.Copy()
        dest_range.PasteSpecial(-4122)
        self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook() 
        return f"copied {source_range} in sheet name {source_sheet_name} to {target_range} in sheet name {target_sheet_name}"
    
    def copy_paste_with_format(self, source_workbook_path: str, source_sheet_name: str, target_sheet_name: str, source_range: str, target_range: str, output_workbook_path: str):
        self.copy_paste(source_workbook_path, source_sheet_name, target_sheet_name, source_range, target_range, output_workbook_path)
        self.copy_paste_format(source_workbook_path, source_sheet_name, target_sheet_name, source_range, target_range, output_workbook_path)
        
            
    def copy_sheet(self, source_workbook_file, source_sheet, target_sheet, output_workbook_path, before = False):
        self.OpenWorkbook(source_workbook_file)
        if source_sheet not in [sheet.Name for sheet in self.activeWB.Worksheets]:
            raise ValueError(f'Sheet {source_sheet} does not exist.')

        source = self.activeWB.Sheets(source_sheet)
        if target_sheet:
            if target_sheet not in [sheet.Name for sheet in self.activeWB.Worksheets]:
                raise ValueError(f'Target sheet {target_sheet} does not exist.')
            target = self.activeWB.Sheets(target_sheet)
            if before:
                source.Copy(Before=target)
            else:
                source.Copy(After=target)
        else:
            source.Copy(After=self.activeWB.Sheets(self.activeWB.Sheets.Count))
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
        return f"copied sheet name {source_sheet} as new sheet {target_sheet} to new workbook"
        
    def cut_paste(self, src_workbook_path, tgt_workbook_path, src_sheet_name,tgt_sheet_name, source_range, target_range):
        self.copy_paste(src_workbook_path, src_sheet_name, tgt_sheet_name, source_range, target_range, tgt_workbook_path)
        self.delete_cells(src_workbook_path, src_sheet_name, source_range, src_workbook_path)
        
        
    def find_n_replace(self, workbook_path, sheet_name, cell_range, find_text, replace_text, output_worbook_path):
        # self.OpenWorkbook(workbook_path)
        sheet = self.activeWB.Sheets(sheet_name)
        range_to_search = self.toRange(sheet, cell_range)
        range_to_search.Replace(find_text, replace_text)
        self.Save()
        self.SaveWorkbook(output_worbook_path)
        self.closeWorkBook()
        return f"found value {replace_text} and replaced it with {find_text} in sheet {sheet_name}"
    
    def set_hyperlink(self, workbook_path, sheet_name, cell_range, url, output_workbook_path):
        # self.OpenWorkbook(workbook_path)
        sheet = self.activeWB.Sheets(sheet_name)
        source = self.toRange(sheet, cell_range)
        sheet = source.Parent
        sheet.Hyperlinks.Add(Anchor=source, Address=url, TextToDisplay=str(source.Value))
        self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
        return f"the hyper link {url} was set for the cell range {cell_range} in sheet {sheet_name}"
    
    def delete_hyperlink(self, workbook_path, sheet_name, cell_range, output_workbook_path):
        # self.OpenWorkbook(workbook_path)
        sheet = self.activeWB.Sheets(sheet_name)
        source = self.toRange(sheet, cell_range)
        source.ClearHyperlinks()
        self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
        return f"the hyper link was removed for the cell range {cell_range} in sheet {sheet_name}"
    
    
    def remove_duplicates(self, workbook_path, sheet_name, column_number, cell_range, target_workbook_path):
        # self.OpenWorkbook(workbook_path)
        sheet = self.activeWB.Sheets(sheet_name)
        source = self.toRange(sheet, cell_range)
        source.RemoveDuplicates(Columns=[column_number], Header=win32.constants.xlNo)  # Columns should be a list
        self.Save()
        self.SaveWorkbook(target_workbook_path)
        self.closeWorkBook()
        return f"Duplicates were removed from column {column_number} in sheet {sheet_name}"

    def rename_sheet(self, workbook_path, old_sheet_name, new_sheet_name, target_workbook_path):
        # self.OpenWorkbook(workbook_path)
        sheet = self.activeWB.Sheets(old_sheet_name)
        sheet.Name = new_sheet_name
        self.Save()
        self.SaveWorkbook(target_workbook_path)
        self.closeWorkBook()
        return f"the sheet {old_sheet_name} was renamed to {new_sheet_name}"
    
    
    def insert_checkbox(self, workbook_path, sheet_name, cell_range, output_workbook_path, width: int = 10, height: int = 10):
        # self.OpenWorkbook(workbook_path)
        sheet = self.activeWB.Sheets(sheet_name)
        sheet.CheckBoxes().Add(Left=sheet.Range(cell_range).Left, Top=sheet.Range(cell_range).Top, Width = width, Height = height)
        self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
        return f"Inserted checkboxes for the range {cell_range} on sheet {sheet_name}"
    
    
    def insert_textbox(self, workbook_path, sheet_name, cell_range, output_workbook_path, text = "", width: int = 10, height: int = 10):
        # self.OpenWorkbook(workbook_path)
        sheet = self.activeWB.Sheets(sheet_name)
        textbox = sheet.Shapes.AddTextbox(1, Left=sheet.Range(cell_range).Left, Top=sheet.Range(cell_range).Top, Width = width, Height =  height)
        textbox.TextFrame.Characters().Text = text
        self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
        return f"Inserted checkboxes for the range {cell_range} on sheet {sheet_name}"
    
    
    def create_sheet(self, workbook_path, sheet_name):
        # self.OpenWorkbook(workbook_path)
        self.activeWB.Sheets.Add(After = self.activeWB.Sheets(self.activeWB.Sheets.Count))
        new_sheet = self.activeWB.Sheets(self.activeWB.Sheets.Count)
        new_sheet.Name = sheet_name
        self.Save()
        self.SaveWorkbook(workbook_path)
        self.closeWorkBook()
        return f"created new sheet with name {sheet_name}"
    
    def delete_sheet(self, workbook_path, sheet_name):
        # self.OpenWorkbook(workbook_path)
        self.activeWB.Sheets.Add(After = self.activeWB.Sheets(self.activeWB.Sheets.Count))
        try:
            sheet = self.activeWB.Sheets(sheet_name)
            sheet.Delete()
        except Exception as e:
            print(f"Error: {e}")
        self.Save()
        self.SaveWorkbook(workbook_path)
        self.closeWorkBook()
        return f"deleted sheet with name {sheet_name}"
    
    def clear(self, workbook_path, sheet_name, cell_range, output_workbook_path):
        # self.OpenWorkbook(workbook_path)
        try:
        # Get the sheet by name
            sheet = self.activeWB.Sheets(sheet_name)
        # Get the range to clear
            range_to_clear = sheet.Range(cell_range)
        # Clear the range
            range_to_clear.ClearContents()
        # Save the workbook
        except Exception as e:
            print(f"Error: {e}")
        self.Save()
#        self.SaveWorkbook(output_workbook_path)
#        self.closeWorkBook()
        return f"deleted sheet with name {sheet_name}"
            
# path = "C:\\Users\\ramy6\\Downloads\\Book1.xlsx"
# app = entry_manipulation_App()
# time.sleep(3)
# app.OpenWorkbook(path)
# time.sleep(3)
# print(app.update_cell_value(path, "Sheet1", 'A2:A10', 674, path))
# time.sleep(3)
# print(app.delete_cells(path, "Sheet1", 'A2:A10', path))
# time.sleep(3)
# print(app.merge_cells(path, "Sheet1", 'D2:D10', path))
# time.sleep(3)
# print(app.unmerge_cells(path, "Sheet1", 'D2:D10', path))
# app.Save()
# app.SaveWorkbook(path)
# app.split_text_to_columns("./DemographicProfile.xlsx", "Sheet1", 'D2:D10', "./DemographicProfile.xlsx", ",") <== works
# app.insert_rows("./DemographicsProfile.xlsx", "Sheet1", 5, 6, 10,"./DemographicProfile.xlsx") <== works
# app.insert_columns("./DemographicProfile.xlsx", "Sheet1", 2, 3, 10,"./DemographicProfile.xlsx") <== works
# app.autofill("./BoomerangSales.xlsx", "Sheet1", 'E2:E10', "Sheet1", 'E2:E10', "./BoomerangSales.xlsx") <== works
# app.copy_paste("./IncomeStatement2.xlsx", "Sheet1", "sheet replica", 'A2:A10', 'G2:G10', "./IncomeStatement2.xlsx") <== works
# app.copy_paste_format("./IncomeStatement2.xlsx", "Sheet1", "sheet replica", 'B1:B10', 'B1:B10', "./IncomeStatement2.xlsx") <== works
# app.copy_sheet("./IncomeStatement2.xlsx", "sheet replica", "Sheet2", "./IncomeStatement2.xlsx") <== works
# app.cut_paste("./IncomeStatement2.xlsx","./IncomeStatement2.xlsx","sheet replica", "Sheet2", "A1:A20", "D1:D20") <== works
# app.find_n_replace("./IncomeStatement2.xlsx", "Sheet1", 'A2:A20', "2033", "work", "./IncomeStatement2.xlsx") <== works
# app.set_hyperlink("./IncomeStatement2.xlsx", "Sheet1", 'A2:A10', "https://chromewebstore.google.com/?hl=en&pli=1", "./IncomeStatement2.xlsx") <== works
# app.delete_hyperlink("./IncomeStatement2.xlsx", "Sheet1", 'A2:A10', "./IncomeStatement2.xlsx") <== works
# app.remove_duplicates("./DemographicProfile.xlsx", "Sheet1", 5, 'E1:E10', "./DemographicProfile.xlsx") <== fix this
# app.rename_sheet("./IncomeStatement2.xlsx", "Sheet2", "sheet replica", "./IncomeStatement2.xlsx") <== works
# app.insert_checkbox("./IncomeStatement2.xlsx", "Sheet2", "A1:A1","./IncomeStatement2.xlsx") <== works
# app.insert_textbox("./IncomeStatement2.xlsx", "Sheet1", "A1:A1","./IncomeStatement2.xlsx", "hello") <== works
# app.create_sheet("./IncomeStatement2.xlsx", "code") <== works
# app.delete_sheet("./IncomeStatement2.xlsx", "code") <== works
# app.clear("./IncomeStatement2.xlsx", "Sheet1", 'G2:G10', "./IncomeStatement2.xlsx") <== works