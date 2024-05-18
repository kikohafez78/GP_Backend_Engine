from xlwings import constants as win32c
from constants import constants
import win32api
import win32com.client as win32
from typing import Any, Optional, List
from openpyxl.utils import get_column_letter
import os
import itertools

class xwBackend():
    def __init__(self, app = 'excel', api_doc = None) -> None:
        super().__init__()
        self.appName = app
        self.__excel = None
        self.__currentWB = None

        if api_doc is not None:
            for key, value in api_doc.items():
                if value.get('display') is not None:
                    setattr(self, value['display'], self.__getattribute__(key))

    @property
    def activeAPP(self):
        if not self.__excel:
            try:
                self.__excel = win32.Dispatch('Excel.Application') if self.appName == 'excel' else win32.Dispatch('ket.Application')
                self.__excel.DisplayAlerts = False
                self.__excel.Visible = False
            except:
                raise Exception('{} is not running.'.format(self.appName))
        return self.__excel
    
    @property
    def activeWB(self):
        if self.__currentWB is not None:
            return self.__currentWB
        return self.activeAPP.ActiveWorkbook
    
    @activeWB.setter
    def activeWB(self, wb):
        self.__currentWB = wb
    
    @property
    def activeWS(self):
        return self.activeWB.ActiveSheet
    
    def toRange(self, source: str):
        if '!' in source:
            sheet_name, source = source.split('!')
            sheet_name = sheet_name.strip("'") # Sheet names with spaces are enclosed with single quotes which should be removed
            # check if the sheet exists
            if sheet_name not in [sheet.Name for sheet in self.activeWB.Worksheets]:
                raise ValueError(f'Sheet {sheet_name} does not exist.')
            sheet = self.activeWB.Sheets(sheet_name)
        else:
            raise Exception('The range must contain a sheet name.')
        
        if source.isdigit():
            return sheet.Rows(source)
        elif source.isalpha():
            return sheet.Columns(source)
        elif ':' in source or source.isalnum():
            return sheet.Range(source)
    
    def OpenWorkbook(self, path: str) -> None:
        self.__currentWB = self.activeAPP.Workbooks.Open(os.path.abspath(path))

    def SaveWorkbook(self, path: str) -> None:
        self.activeWB.SaveAs(os.path.abspath(path))
                
    def Write(self, range: str, value: Any) -> None:
        range = self.toRange(range)
        
        if range.Rows.Count > 5000:
            raise Exception('The range is too large. Please note that reference to whole columns like A:A and C:D actually includes millions of rows. Please use a specific range like A1:C30 instead.')
        
        if isinstance(value, (list, tuple)) and range.Count == 1:
            if isinstance(value[0], (list, tuple)):
                for rowOffet, elem in enumerate(value):
                    for columnOffset, elem2 in enumerate(elem):
                        range.GetOffset(rowOffet, columnOffset).Value = elem2
            else:
                for columnOffset, elem in enumerate(value):
                    range.GetOffset(0, columnOffset).Value = elem
        else:
            range.Value = value

    def CopyPaste(self, source: str, destination: str) -> None:
        source = self.toRange(source).SpecialCells(12)
        destination = self.toRange(destination)
        source.Copy()
        destination.PasteSpecial(-4163)
        # self.CopyPasteVisible(source, destination)

    def CopyPasteVisible(self, source: str, destination: str) -> None:
        source = self.toRange(source)
        rowCount, columnCount = source.Rows.Count, source.Columns.Count
        firstCell = source.Cells(1,1)
        while firstCell.EntireRow.Hidden:
            firstCell = firstCell.GetOffset(1,0)
        while firstCell.EntireColumn.Hidden:
            firstCell = firstCell.GetOffset(0,1)
        lastCell = firstCell
        while rowCount > 0:
            lastCell = lastCell.GetOffset(1,0)
            if not lastCell.EntireRow.Hidden:
                rowCount -= 1
        while columnCount > 0:
            lastCell = lastCell.GetOffset(0,1)
            if not lastCell.EntireColumn.Hidden:
                columnCount -= 1
        destination = self.toRange(destination)
        source.Copy()
        destination.PasteSpecial(-4163)
    
    def CopyPasteFormat(self, source: str, destination: str) -> None:
        source = self.toRange(source)
        destination = self.toRange(destination)
        source.Copy()
        destination.PasteSpecial(-4122)

    def CutPaste(self, source: str, destination: str) -> None:
        source = self.toRange(source)
        destination = self.toRange(destination)
        source.Cut(Destination=destination)

    def FindReplace(self, source: str, find: str, replace: str) -> None:
        source = self.toRange(source)
        source.Replace(find, replace)

    def SetHyperlink(self, source: str, url: str) -> None:
        source = self.toRange(source)
        sheet = source.Parent
        sheet.Hyperlinks.Add(Anchor=source, Address=url, TextToDisplay=str(source.Value))

    def RemoveHyperlink(self, source: str) -> None:
        source = self.toRange(source)
        source.ClearHyperlinks()

    def RenameSheet(self, oldName: str, newName: str) -> None:
        self.activeWB.Sheets(oldName).Name = newName

    def WrapText(self, range: str) -> None:
        range = self.toRange(range)
        range.WrapText = True

    def UnwrapText(self, range: str) -> None:
        range = self.toRange(range)
        range.api.WrapText = False

    def AutoFill(self, source: str, destination: str) -> None:
        source = self.toRange(source)
        destination = self.toRange(destination)
        
        # source_row_start, source_row_end, source_column_start, source_column_end = source.Row, source.Row + source.Rows.Count, source.Column, source.Column + source.Columns.Count
        # destination_row_start, destination_row_end, destination_column_start, destination_column_end = destination.Row, destination.Row + destination.Rows.Count, destination.Column, destination.Column + destination.Columns.Count
        if self.activeAPP.Union(destination, source).Address != destination.Address:
            raise ValueError('Illegal source and destination! The auto-filling destination must include the source!')

        source.AutoFill(destination)

    def Sort(self, source: str, key1: str, order: str='asc', orientation: str='column') -> None:
        source = self.toRange(source)
        key1 = self.toRange(key1)
        source.Sort(Key1=key1, Order1=1 if order == 'asc' else 2, 
                        Orientation=1 if orientation == 'column' else 2)

    def Filter(self, source: str, fieldIndex: int, criteria: str) -> None:
        source = self.toRange(source)
        try:
            criteriaRange = self.toRange(criteria)
        except:
            criteriaRange = None
        if criteriaRange:
            criteria = [criteriaRange.Cells(i).Text for i in range(1, criteriaRange.Cells.Count + 1)]
        source.AutoFilter(Field=fieldIndex, Criteria1=criteria)
        # source.AutoFilter(Field=field, Criteria1=criteria, Operator=constants.AutoFilterOperator['values'])

    def DeleteFilter(self) -> None:
        if self.activeWS.AutoFilterMode: self.activeWS.AutoFilterMode = False

    def MoveRow(self, source: int, destination: int) -> None:
        if destination > source:
            destination += 1
        else:
            source += 1
        self.InsertRow(destination, aboveRow=destination)
        self.activeWS.Rows(source).Copy(self.activeWS.Rows(destination))
        self.activeWS.Rows(source).Delete()

    def MoveColumn(self, source: int, destination: int) -> None:
        if destination > source:
            destination += 1
        else:
            source += 1
        self.InsertColumn(destination)
        self.activeWS.Columns(source).Copy(self.activeWS.Columns(destination))
        self.activeWS.Columns(source).Delete()
    
    def RemoveDuplicate(self, source: str, key: int) -> None:
        source = self.toRange(source)
        source.RemoveDuplicates(key)

    def group_ungroup(self) -> None:
        pass

    def SetPassword(self, password: str) -> None:
        self.activeWB.Password = password

    def TransposeRange(self, source: str) -> None:
        source = self.toRange(source)
        dataT = self.activeAPP.WorksheetFunction.Transpose(source)
        source.Clear()
        cell = source.Cells(1).Address
        self.Write(cell, dataT)

    def CreateNamedRange(self, source: str, name: str):
        source = self.toRange(source)
        source.Name = name

    def SetFormat(self, source: str, font: Optional[str] = None, fontSize: Optional[float] = None,
                    color: Optional[int] = None, fillColor: Optional[int] = None, bold: Optional[bool] = None,
                    italic: Optional[bool] = None, underline: Optional[bool] = None, horizontalAlignment: Optional[str] = None) -> None:
        source = self.toRange(source)
        if font:
            source.Font.Name = font
        if fontSize:
            source.Font.Size = fontSize
        if color:
            source.Font.ColorIndex = constants.ColorIndex[color]
        if fillColor:
            source.Interior.ColorIndex = constants.ColorIndex[fillColor]
        if not bold is None:
            source.Font.Bold = bold
        if not italic is None:
            source.Font.Italic = italic
        if not underline is None:
            source.Font.Underline = win32c.UnderlineStyle.xlUnderlineStyleSingle if underline else win32c.UnderlineStyle.xlUnderlineStyleNone
        if horizontalAlignment:
            source.HorizontalAlignment = constants.HorizontalAlignment[horizontalAlignment]
    
    def DeleteFormat(self, source: str) -> None:
        source = self.toRange(source)
        source.ClearFormats()

    def SetDataType(self, source: str, dataType: str) -> None:
        source = self.toRange(source)
        source.NumberFormat = constants.DataType[dataType]

    def SetPageLayout(self, orientation: str, paperSize: str) -> None:
        self.activeWS.PageSetup.Orientation = constants.PageOrientation[orientation]
        self.activeWS.PageSetup.PaperSize = constants.PaperSize[paperSize]

    def SetBorderAround(self, source: str, color: str, weight: str) -> None:
        source = self.toRange(source)
        source.BorderAround(ColorIndex=constants.ColorIndex[color], Weight=constants.BorderWeight[weight])

    def ToggleRowColumnVisibility(self, range: str, visible: bool, region: str) -> None:
        range = self.toRange(range)
        if region == 'row':
            range.EntireRow.Hidden = not visible
        elif region == 'column':
            range.EntireColumn.Hidden = not visible

    def SetCellMerge(self, source: str, merge: bool) -> None:
        source = self.toRange(source)
        if merge:
            source.Merge()
        else:
            source.Unmerge()

    def merging_text(self) -> None:
        pass

    def Delete(self, source: str, region: str) -> None:
        source = self.toRange(source)
        if region == 'row':
            source.EntireRow.Delete()
        elif region == 'column':
            source.EntireColumn.Delete()
        else:
            source.Delete()

    def Clear(self, source: str) -> None:
        source = self.toRange(source)
        source.Clear()

    def Insert(self, source: str, shift: Optional[str]) -> None:
        source = self.toRange(source)
        source.Insert(constants.InsertShiftDirection[shift])

    def InsertRow(self, sheetName: str, aboveRow: int = None, belowRow: int = None) -> None:
        if aboveRow:
            index = aboveRow
        elif belowRow:
            index = belowRow + 1
        sheet = self.activeWB.Sheets(sheetName)
        lastCell = sheet.UsedRange(sheet.UsedRange.Count)
        source = sheet.Range(sheet.Cells(index, 1), sheet.Cells(index, lastCell.Column))
        source.Insert(constants.InsertShiftDirection['down'])
    
    def InsertColumn(self, sheetName: str, beforeColumn: str = None, afterColumn: str = None) -> None:
        if beforeColumn:
            index = self.activeWS.Columns(beforeColumn).Column
        elif afterColumn:
            index = self.activeWS.Columns(afterColumn).Column + 1
        sheet = self.activeWB.Sheets(sheetName)
        lastCell = sheet.UsedRange(sheet.UsedRange.Count)
        source = sheet.Range(sheet.Cells(1, index), sheet.Cells(lastCell.Row, index))
        source.Insert(constants.InsertShiftDirection['right'])

    def SplitText(self, source: str, delimiter: str) -> None:
        source = self.toRange(source)
        if source.Columns.Count != 1:
            print('Source must be one column.')
            return
        row = source.Row
        column = source.Column
        orgData = source.Value
        newData = [x.split(delimiter) for x in orgData]
        maxLen = max(len(x) for x in newData)
        newData = [x + [None] * (maxLen-len(x)) for x in newData]
        for i in range(maxLen - 1):
            self.InsertColumn(source.Column)
        self.Write(self.activeWS.Cells(row,column).Address, newData)

    def AutoFit(self, source: str) -> None:
        source = self.toRange(source)
        source.AutoFit()

    def ResizeRowColumn(self, source: str, height: Optional[int] = None, width: Optional[int] = None) -> None:
        source = self.toRange(source)
        if height:
            source.RowHeight = height
        if width:
            source.ColumnWidth = width

    def SetConditionalFormat(self, source: str, formula: str,
                            bold: Optional[bool] = None, color: Optional[str] = None,
                            fillColor: Optional[str] = None, italic: Optional[bool] = None, underline: Optional[bool] = None) -> None:
        source = self.toRange(source)

        handle = source.FormatConditions.Add(Type = constants.FormatConditionType['expression'], Formula1 = formula)
        if color:
            handle.Font.ColorIndex = constants.ColorIndex[color.lower()]
        if fillColor:
            handle.Interior.ColorIndex = constants.ColorIndex[fillColor.lower()]
        if not bold is None:
            handle.Font.Bold = bold
        if not italic is None:
            handle.Font.Italic = italic
        if not underline is None:
            handle.Font.Underline = win32c.UnderlineStyle.xlUnderlineStyleSingle if underline else win32c.UnderlineStyle.xlUnderlineStyleNone

    def SetDataValidation(self, source: str, type: str, formula1: str) -> None:
        source = self.toRange(source)
        handle = source.Validation.Add(constants.ValidationType[type], Formula1 = formula1)

    def ToggleFormulaDisplay(self, display: bool) -> None:
        self.activeAPP.ActiveWindow.DisplayFormulas = display

    def SplitPanes(self, rowCount: int, columnCount: int) -> None:
        self.activeAPP.ActiveWindow.SplitRow = rowCount
        self.activeAPP.ActiveWindow.SplitColumn = columnCount

    def SetCellLock(self, source: str, lock: bool) -> None:
        source = self.toRange(source)
        source.Locked = lock

    def ToggleSheetProtection(self, sheetName: str, protect: bool, password: str = None) -> None:
        sheet = self.activeWB.Worksheets(sheetName)
        if protect:
            sheet.Protect(password)
        else:
            sheet.Unprotect(password)

    def FreezePanes(self, source: str) -> None:
        source = self.toRange(source)
        source.Select()
        source.Parent.Application.ActiveWindow.FreezePanes = True

    def UnfreezePanes(self, sheetName: Optional[str] = None) -> None:
        if sheetName:
            sheet = self.activeWB.Worksheets(sheetName)
        else:
            sheet = self.activeWS
        sheet.Application.ActiveWindow.FreezePanes = False

    def CreateChart(self, source: str, destSheet: str, chartType: str, chartName: str, XField: int = None, YField: List[int] = []) -> None:
        # check if the chart name exists
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    raise ValueError(f'The chart name {chartName} already exists.')
        # check if the destSheet exists
        if destSheet not in [sheet.Name for sheet in self.activeWB.Worksheets]:
            raise ValueError(f'Sheet {destSheet} does not exist.')
        dataRange = self.toRange(source)
        destRange = self.GetBlankArea(destSheet)
        sheet = self.activeWB.Worksheets(destSheet)
        chart = sheet.ChartObjects().Add(destRange.Left, destRange.Top, 350, 200).Chart
        
        if chartType not in constants.ChartType:
            raise ValueError(f'Chart type {chartType} is not supported!')
        
        chart.ChartType = constants.ChartType[chartType]
        
        if 'pie' in chartType.lower():
            chart.SetSourceData(dataRange)
        else:
            if not XField:
                XField = 1
            XFieldRange = dataRange.Parent.Range(dataRange.Cells(2, XField), dataRange.Cells(dataRange.Rows.Count, XField))
            if not YField:
                YField = [i for i in range(1, dataRange.Columns.Count + 1) if i != XField]
            for i in YField:
                series = chart.SeriesCollection().NewSeries()
                series.XValues = XFieldRange
                series.Values = dataRange.Parent.Range(dataRange.Cells(2, i), dataRange.Cells(dataRange.Rows.Count, i))
                series.Name = dataRange.Cells(1, i)
            try:
                chart.Axes(constants.AxisType['x']).CategoryNames = XFieldRange
            except:
                pass
        chart.Parent.Name = chartName

    def SetChartTrendline(self, chartName: str, trendlineType: List[str], DisplayEquation: Optional[bool] = None,
                          DisplayRSquared: Optional[bool] = None) -> None:
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        trendlineType = itertools.cycle(trendlineType)
        for series in chart.SeriesCollection():
            for trendline in series.Trendlines():
                trendline.Delete()
            series.Trendlines().Add(
                constants.TrendlineType[next(trendlineType)],
                DisplayEquation=DisplayEquation,
                DisplayRSquared=DisplayRSquared
            )

    def SetChartTitle(self, chartName: str, title: str, fontSize: Optional[float] = None, 
                        bold: bool = None, color: Optional[int] = None) -> None:
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        chart.HasTitle = True
        chart.ChartTitle.Text = title
        if fontSize:
            chart.ChartTitle.Font.Size = fontSize
        if color:
            chart.ChartTitle.Font.ColorIndex = {
                'black': 1,
                'white': 2,
                'red': 3,
                'green': 4,
                'blue': 5,
                'yellow': 6,
                'magenta': 7,
                'cyan': 8,
                'dark red': 9,
                'dark green': 10
            }[color]
        if not bold is None:
            chart.ChartTitle.Font.Bold = bold

    def SetChartHasAxis(self, chartName: str, axis: str, hasAxis: bool) -> None:
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        axis = constants.AxisType[axis]
        axisGroup = constants.AxisGroup['primary']
        chart.SetHasAxis(axis, axisGroup, hasAxis)
        

    def SetChartAxis(self, chartName: str, axis: str, title: Optional[str] = None, 
                        labelOrientation: Optional[str] = None, maxValue: Optional[float] = None,
                        miniValue: Optional[float] = None) -> None:
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        if axis in ['x', 'X']:
            axis = win32c.AxisType.xlCategory
        elif axis in ['y', 'Y']:
            axis = win32c.AxisType.xlValue
        elif axis in ['z', 'Z']:
            axis = win32c.AxisType.xlSeriesAxis
        else:
            print('Not support axes type')
            return
        chartAxes = chart.Axes(axis)
        if title:
            chartAxes.HasTitle = True
            chartAxes.AxisTitle.Text = title
        if labelOrientation:
            labelOrientation = {
                'upward': win32c.Orientation.xlUpward,
                'downward': win32c.Orientation.xlDownward,
                'horizontal': win32c.Orientation.xlHorizontal,
                'vertical': win32c.Orientation.xlVertical
                }[labelOrientation]
            chartAxes.TickLabels.Orientation = labelOrientation
        if maxValue:
            chartAxes.MaximumScale = maxValue
        if miniValue:
            chartAxes.MinimumScale = miniValue

    def SetChartLegend(self, chartName: str, position: Optional[str] = None, fontSize: Optional[str] = None,
                        seriesName: Optional[list] = []) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        chart.HasLegend = True
        
        if position and position != 'None':
            # For legent position enumeration, refer to https://learn.microsoft.com/en-us/dotnet/api/microsoft.office.interop.word.xllegendposition?view=word-pia
            position = {
                'bottom': win32c.LegendPosition.xlLegendPositionBottom,
                'corner': win32c.LegendPosition.xlLegendPositionCorner,
                'left': win32c.LegendPosition.xlLegendPositionLeft,
                'right': win32c.LegendPosition.xlLegendPositionRight,
                'top': win32c.LegendPosition.xlLegendPositionTop
            }[position]
            chart.Legend.Position = position
        if seriesName:
            for index, elem in enumerate(seriesName):
                chart.SeriesCollection(index+1).Name = elem

    def SetChartHasLegend(self, chartName: str, hasLegend: bool) -> None:
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        chart.HasLegend = hasLegend

    def SetChartType(self, chartName: str, chartType: str) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        chart.ChartType = constants.ChartType[chartType]

    def SetChartSource(self, chartName: str, source: str) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        source = self.toRange(source)
        chart.SetSourceData(source)

    def SetChartBackgroundColor(self, chartName: str, color: str) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        chart.ChartArea.Interior.ColorIndex = constants.ColorIndex[color]

    def ResizeChart(self, chartName: str, width: float, height: float) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart.Width = width
        chart.Height = height

    def SetChartDataColor(self, chartName: str, colorRGB: list) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        for series in chart.SeriesCollection():
            for point in series.Points():
                point.Format.Fill.ForeColor.RGB = win32api.RGB(*colorRGB)
    
    def HighlightDataPoints(self, chartName: str, pointIndex: int, colorRGB: list) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        series = chart.SeriesCollection(1)
        point = series.Points(pointIndex)
        point.Format.Fill.ForeColor.RGB = win32api.RGB(*colorRGB)

    def SetDataSeriesType(self, chartName: str, seriesIndex: int, seriesType: str) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        series = chart.SeriesCollection(seriesIndex)
        series.ChartType = constants.ChartType[seriesType]

    def AddDataSeries(self, chartName: str, xrange: str, yrange: str) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        xrange = self.toRange(xrange)
        yrange = self.toRange(yrange)
        series = chart.SeriesCollection().NewSeries()
        series.XValues = xrange
        series.Values = yrange
    
    def RemoveDataSeries(self, chartName: str, seriesIndex: int) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        series = chart.SeriesCollection(seriesIndex)
        series.Delete()

    def SetDataSeriesSource(self, chartName: str, seriesIndex: int, xrange: str, yrange: str) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        xrange = self.toRange(xrange)
        yrange = self.toRange(yrange)
        series = chart.SeriesCollection(seriesIndex)
        series.XValues = xrange
        series.Values = yrange

    def AddChartErrorBars(self, chartName: str) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        for series in chart.SeriesCollection():
            series.HasErrorBars = True

    def AddChartErrorBar(self, chartName: str, seriesIndex: int) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        series = chart.SeriesCollection(seriesIndex)
        series.HasErrorBars = True

    def RemoveChartErrorBars(self, chartName: str) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        for series in chart.SeriesCollection():
            series.HasErrorBars = False

    def RemoveChartErrorBar(self, chartName: str, seriesIndex: int) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        series = chart.SeriesCollection(seriesIndex)
        series.HasErrorBars = False

    def AddDataLabels(self, chartName: str) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        for series in chart.SeriesCollection():
            series.HasDataLabels = True
    
    def RemoveDataLabels(self, chartName: str) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        for series in chart.SeriesCollection():
            series.HasDataLabels = False

    def SetChartMarker(self, chartName: str, style: List[str] = None, size: Optional[float] = None) -> None:
        '''
        style: auto, circle, dash, dot, star, triangle, square, plus
        '''
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        styleIter = itertools.cycle(style)
        for series in chart.SeriesCollection():
            if style:
                series.MarkerStyle = constants.MarkerStyle[next(styleIter)]
            if size:
                series.MarkerSize = size

    def CopyPasteChart(self, chartName, destination: str) -> None:
        # find the chart
        chart = None
        for sheet in self.activeWB.Worksheets:
            for chart in sheet.ChartObjects():
                if chart.Name == chartName:
                    break
        if chart is None or chart.Name != chartName:
            raise ValueError(f'Chart {chartName} does not exist.')
        chart = chart.Chart
        destination = self.toRange(destination)
        chart.ChartArea.Copy()
        destination.Select()
        destination.Parent.Paste()

    def CreatePivotTable(self, source: str, destSheet: str, name: str,
                        RowField: List = [], ColumnField: List = [],
                        PageField: List = [], DataField: List = [],
                        summarizeFunction = 'sum') -> None:
        # check if the pivot table name exists
        for sheet in self.activeWB.Worksheets:
            for pt in sheet.PivotTables():
                if pt.Name == name:
                    raise ValueError(f'Pivot table {name} already exists. Please choose a different name.')
        # check if the destSheet exists
        if destSheet not in [sheet.Name for sheet in self.activeWB.Worksheets]:
            raise ValueError(f'Sheet {destSheet} does not exist.')
        
        # # check if the four fields are letters
        # invalid_fields = []
        # if any([len(x) > 1 for x in RowField]):
        #     invalid_fields.append('RowField')
        # if any([len(x) > 1 for x in ColumnField]):
        #     invalid_fields.append('ColumnField')
        # if any([len(x) > 1 for x in PageField]):
        #     invalid_fields.append('PageField')
        # if any([len(x) > 1 for x in DataField]):
        #     invalid_fields.append('DataField')
        
        # if len(invalid_fields) > 0:
        #     raise ValueError('Illegal fields! the fields in {} can only be column indices (i.e., letters A to Z)'.format(",".join(invalid_fields)))
        
        # sheet = self.activeWB.Worksheets(destSheet)
        sourceRange = self.toRange(source)
        sourceRange_sheet = sourceRange.Worksheet
        # Sometimes the LLM misses the header row, so we manually add it
        if sourceRange.Row != 1:
            new_starting_cell = sourceRange_sheet.Cells(1, sourceRange.Column)
            sourceRange = sourceRange_sheet.Range(new_starting_cell, sourceRange_sheet.Cells(new_starting_cell.Row + sourceRange.Rows.Count, sourceRange.Column + sourceRange.Columns.Count - 1))
            
        pc = self.activeWB.PivotCaches().Create(SourceType=win32c.PivotTableSourceType.xlDatabase, SourceData=sourceRange)
        destRange = self.GetBlankArea(destSheet)

        pt = pc.CreatePivotTable(TableDestination=destRange, TableName=name)
        for field in RowField:
            try:
                pt.PivotFields(field).Orientation = win32c.PivotFieldOrientation.xlRowField
            except:
                raise Exception(f"The field {field} is not included in the pivot table source range and cannot be selected as a row field!")
        for field in ColumnField:
            try:
                pt.PivotFields(field).Orientation = win32c.PivotFieldOrientation.xlColumnField
            except:
                raise Exception(f"The field {field} is not included in the pivot table source range and cannot be selected as a column field!")
        # pc.CreatePivotTable(TableDestination=destRange, TableName=name)
        # pt = sheet.PivotTables(name)

        for field in PageField:
            try:
                pt.PivotFields(field).Orientation = win32c.PivotFieldOrientation.xlPageField
            except:
                raise Exception(f"The field {field} is not included in the pivot table source range and cannot be selected as a page field!")
        
        for field in DataField:
            try:
                pt.AddDataField(pt.PivotFields(field), Function = constants.SummarizationFunction[summarizeFunction])
            except:
                raise Exception(f"The field {field} is not included in the pivot table source range and cannot be selected as a data field!")
            #pt.PivotFields(field).Orientation = win32c.PivotFieldOrientation.xlDataField
            # pt.PivotFields(field).Function = constants.ConsolidationFunction[summarizeFunction]

    def GetBlankArea(self, sheetName: str):
        sheet = self.activeWB.Sheets(sheetName)
        chartRangeList = []
        for chart in sheet.ChartObjects():
            chartRangeList.append(sheet.Range(chart.TopLeftCell, chart.BottomRightCell))
        row, column = 1, 1
        checkScope = 5
        while True:
            cell1 = sheet.Cells(row,column)
            cell2 = sheet.Cells(row+checkScope,column+checkScope)
            checkRange = sheet.Range(cell1, cell2)
            if all(cell.Value is None for cell in checkRange) and all(self.activeAPP.Intersect(chartRange, checkRange) is None for chartRange in chartRangeList):
                break
            row += 1
            column += 1

        while row > 1:
            cell1 = sheet.Cells(row-1,column)
            cell2 = sheet.Cells(row-1,column+checkScope)
            checkRange = sheet.Range(cell1, cell2)
            if any(cell.Value is not None for cell in checkRange) or any(self.activeAPP.Intersect(chartRange, checkRange) is not None for chartRange in chartRangeList):
                break
            row -= 1
            
        while column > 1:
            cell1 = sheet.Cells(row,column-1)
            cell2 = sheet.Cells(row+checkScope,column-1)
            checkRange = sheet.Range(cell1, cell2)
            if any(cell.Value is not None for cell in checkRange) or any(self.activeAPP.Intersect(chartRange, checkRange) is not None for chartRange in chartRangeList):
                break
            column -= 1

        return sheet.Cells(row+1,column+1)

    def CreateChartFromPivotTable(self, pivotTableName: str, destSheet: str, chartName: str, chartType: str) -> None:
        for sheet in self.activeWB.Worksheets:
            pt_name = None
            for pt in sheet.PivotTables():
                pt_name = pt.Name
                print(pt_name, '|', pivotTableName)
                if pt_name == pivotTableName:
                    break
            if pt_name is not None: break
        else:
            pt = None

        if pt is None:
            raise ValueError(f'Pivot table {pivotTableName} does not exist. Note that this API is only for creating chart from data in pivot table.')
        # check if the destSheet exists
        if destSheet not in [sheet.Name for sheet in self.activeWB.Worksheets]:
            raise ValueError(f'Sheet {destSheet} does not exist.')
        sourceRange = pt.TableRange2
        destRange = self.GetBlankArea(destSheet)
        sheet = self.activeWB.Worksheets(destSheet)
        chart = sheet.ChartObjects().Add(destRange.Left, destRange.Top, 350, 200).Chart
        chart.ChartType = constants.ChartType[chartType]
        chart.SetSourceData(sourceRange)
        chart.Parent.Name = chartName
        
    def RemovePivotTable(self, name: str) -> None:
        # find the pivot table
        pt = None
        for sheet in self.activeWB.Worksheets:
            for pt in sheet.PivotTables():
                if pt.Name == name:
                    break
        if pt is None or pt.Name != name:
            raise ValueError(f'Pivot table {name} does not exist.')
        pt.TableRange2.Clear()

    def SetPivotTableSummaryFunction(self, name: str, field: str, func: str) -> None:
        pt = self.activeWS.PivotTables(name)
        pt.PivotFields(field).Function = constants.ConsolidationFunction[func]

    def SortPivotTable(self, name: str, field: str, key: str, oder: str = 'ascending') -> None:
        pt = self.activeWS.PivotTables(name)
        oder = constants.SortOrder[oder]
        pt.PivotFields(field).AutoSort(oder, key)

    def CreateSheet(self, sheetName: Optional[str] = None, before: Optional[str] = None, after: Optional[str] = None) -> None:
        for sheet in self.activeWB.Worksheets:
            if sheet.Name == sheetName:
                raise ValueError(f'Sheet {sheetName} already exists.')
        activeSheet = self.activeWS
        if before is not None:
            beforeSheet = self.activeWB.Worksheets(before)
            newSheet = self.activeWB.Worksheets.Add(Before=beforeSheet)
        elif after is not None:
            afterSheet = self.activeWB.Worksheets(after)
            newSheet = self.activeWB.Worksheets.Add(After=afterSheet)
        else:
            newSheet = self.activeWB.Worksheets.Add()
        if sheetName is not None:
            newSheet.Name = sheetName
        activeSheet.Activate()

    def RemoveSheet(self, sheetName: Optional[str] = None) -> None:

        if not sheetName:
            sheetName = self.activeWS.Name
        self.activeWB.Sheets(sheetName).Delete()

    def SwitchSheet(self, sheetName: str) -> None:
        self.activeWB.Sheets(sheetName).Activate()

    def GetSheetsState(self) -> str:
        states = []
        for ws in self.activeWB.Worksheets:
            if ws.Range('A1').Value is None:
                cell_state = "Sheet \"{}\" {} has no content".format(ws.Name, '(active)' if ws.Name == self.activeWS.Name else '')
            else:
                NumberOfRows = ws.UsedRange.Rows.Count
                NumberOfColumns = ws.UsedRange.Columns.Count
                headers = ws.Range('A1', ws.Cells(1,NumberOfColumns)).Value
                if isinstance(headers, tuple):
                    headers = headers[0]
                else:
                    headers = [headers]
                headers = {get_column_letter(i): header for i, header in enumerate(headers, start=1)}
                
                cell_state = 'Sheet \"{}\" has {} columns (Headers are {}) and {} rows (1 header row and {} data rows)'.format(ws.Name, NumberOfColumns, ', '.join(["{}: \"{}\"".format(col_letter, header) for col_letter, header in headers.items()]), NumberOfRows, NumberOfRows-1)
            
            # Add chart descriptions
            # Iterate through the shapes to find chart objects
            chart_names = []
            for shape in ws.Shapes:
                if shape.HasChart:
                    chart = shape.Chart
                    chart_name = chart.Name
                    chart_names.append(chart_name[chart_name.find(' ')+1:])
                    
            chartNameString = ' and this sheet has the charts whose names are "{}"'.format('", "'.join(chart_names)) if len(chart_names) > 0 else ''
            
            # Iterate through the pivot tables and print their names
            pt_names = []
            for pivot_table in ws.PivotTables():
                pt_names.append(pivot_table.Name)

            if len(pt_names) > 0:
                ptNameString = ' the pivot tables whose names are "{}"'.format('", "'.join(pt_names))
                if len(chart_names) == 0:
                    ptNameString = ' and this sheet has' + ptNameString
                else:
                    ptNameString = ' and' + ptNameString
            else:
                ptNameString = ''
            
            states.append("{}{}{}.".format(cell_state, chartNameString, ptNameString))
                                                                                                                       
        return "Sheet state: " + ' '.join(states)


